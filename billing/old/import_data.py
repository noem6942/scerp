# process.py
from datetime import date, datetime
from openpyxl import load_workbook

from core.models import Tenant
from scerp.mixins import get_admin
from .models import Counter, Subscription


class Import(object):

    def __init__(self, file_path, **kwargs):
        # Init
        self.file_path = file_path
        self.kwargs = kwargs

    @staticmethod
    def read_excel_file(file_path):
        '''read an excel sheet and interprete EVERY cell as string.
            i.e. empty cell -> '
                 111.11 -> '111.11'
                 012 -> '012'
        '''
        # Load the workbook
        wb = load_workbook(filename=file_path, data_only=False)  # data_only=False to get formulas too
        ws = wb.active  # Use the active sheet
        
        # Iterate through the rows in the worksheet
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert each cell to string while keeping leading zeros
            string_row = [
                str(cell).strip() if cell is not None else ''
                for cell in row]
            rows.append(string_row)
        
        return rows

    @staticmethod
    def convert_value(
            value, target_type, date_format='%d.%m.%Y', set_null=True):
        # check None    
        if not value and set_null:
            return None
            
        # Convert    
        try:
            if target_type == date:
                return datetime.strptime(value, date_format)            
            return target_type(value)
        except (ValueError, TypeError) as e:
            raise TypeError(f"{value}: Conversion error '{e}'")
            


class Gunzgen(Import):
    CODE = 'gunzgen'
    DATE_FORMAT_1 = '%d.%m.%Y'
    DATE_FORMAT_2 = '%Y-%m-%d %H:%M:%S'
    SET_NULL = True
    STATUS_TRANSFORM = {
        'Lager': Counter.STATUS.STOCK,
        'Montiert': Counter.STATUS.MOUNTED,
        'Weggeworfen': Counter.STATUS.DISPOSED
    }

    def __init__(self, *args, **kwargs):
        self.tenant = Tenant.objects.get(code=self.CODE)
        super().__init__(*args, **kwargs)

    def load_counters(self):        
        # Init        
        keys = [
            # key: type 
            {'nr': str},
            {None: None},
            {'function': str},
            {'zw': int},
            {'jg': int},
            {None: None},
            {'st': str},
            {None: None},
            {'currency': str},
            {'size': str},
            {None: None},
            {'type': str},
            {'calibration_date': date},
            {'status': str}
        ]        
        admin = get_admin()
        rows = self.read_excel_file(self.file_path)
        
        # Parse
        count = 0
        for row in rows:            
            data = {}

            # check validity
            if 'erstellt durch' in row[0]:
                continue
            elif 'Werk-Nr.' in row[0]:
                continue
            elif 'Zähler-Nr.' in row[0]:
                continue
            elif 'Anzahl Sätze:' in row[0]:
                continue
            elif not row[2]:
                continue
            elif not row[0]:
                break        

            # Get data
            data = {
                'created_by': admin, 
                'modified_by': admin,
                'tenant': self.tenant
            }
            for i, key_dict in enumerate(keys):
                key, type_ = next(x for x in key_dict.items())
                if key:
                    # convert
                    try:
                        data[key] = self.convert_value(
                            row[i], type_, date_format=self.DATE_FORMAT_1,
                            set_null=self.SET_NULL)
                    except Exception as e:
                        # try DATE_FORMAT_2
                        try:
                            data[key] = self.convert_value(
                                row[i], type_, date_format=self.DATE_FORMAT_2,
                                set_null=self.SET_NULL)
                        except Exception as e:                    
                            print(f"Error converting value for key '{key}' with type '{type_}' at row index {i}.")
                            print(f"Row attempted: {row}")
                            print(f"Error: {e}")
                            return
                    
                    # handle specials                    
                    if key == 'function':
                        data[key] = data[key][0]         
                    elif key == 'status':
                        data[key] = self.STATUS_TRANSFORM[data[key]]

            # Store                
            _obj, _created = Counter.objects.update_or_create(
                nr=data.pop('nr'), defaults=data)
            count += 1
            
        return count

    def load_subscriptions(self):   
        keys = [
            {'abo_nr': str},  # Subscription Number
            {'r_empf': str},  # Invoice Recipient
            {'pers_nr': int},  # Personal Number
            {'name_vorname': str},  # Name + First Name
            {None: None},  # None
            {'strasse': str},  # Street
            {'plz_ort': str},  # Postal Code + City
            {None: None},  # None
            {'tarif': int},  # Tariff
            {'periode': int},  # Period
            {'tarif_bez': str},  # Tariff Description
            {'basis': float},  # Base Amount
            {'ansatz_nr': int},  # Approach Number
            {'ansatz': float},  # Approach
            {'tage': int},  # Days
            {'betrag': float},  # Amount
            {'inkl_mwst': float},  # Including VAT
            {'steuercode_zaehler': str},  # Tax Code Meter
            {'berechnungs_code_zaehler': str},  # Calculation Code Meter
            {'steuercode_gebuehren': str},  # Tax Code Fees
            {'berechnungs_code_gebuehren': str},  # Calculation Code Fees
            {'gebuehrentext': str},  # Fee Text
            {'gebuehren_zusatztext': str},  # Additional Fee Text
        ]

        admin = get_admin()
        rows = self.read_excel_file(self.file_path)
        
        # Parse
        count = 0
        for row in rows:            
            data = {}

            # check validity
            if 'Abonnenten Gebührenliste' in row[0]:
                continue
            elif 'AboNr' in row[0]:
                continue
            elif not row[0]:
                break        

            # Get data
            data = {
                'created_by': admin, 
                'modified_by': admin,
                'tenant': self.tenant
            }
            for i, key_dict in enumerate(keys):
                key, type_ = next(x for x in key_dict.items())
                if key:
                    # convert
                    try:
                        data[key] = self.convert_value(
                            row[i], type_, set_null=self.SET_NULL)
                    except Exception as e:                
                        print(f"Error converting value for key '{key}' with type '{type_}' at row index {i}.")
                        print(f"Row attempted: {row}")
                        print(f"Error: {e}")
                        return

            # Store           
            _obj, _created = Subscription.objects.update_or_create(
                abo_nr=data.pop('abo_nr'), 
                r_empf=data.pop('r_empf'),
                pers_nr=data.pop('pers_nr'), 
                tarif=data.pop('tarif'), 
                ansatz_nr=data.pop('ansatz_nr'), 
                defaults=data)            
            count += 1
            
        return count
 