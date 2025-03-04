'''
billing/gesoft_counter_data_import.py
'''
import json
from openpyxl import load_workbook
from pathlib import Path

from django.conf import settings


# Helpers
def write(data, filename):
    file_path = Path(settings.BASE_DIR) / "billing" / "fixtures" / filename
    with open(file_path, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
    print("*writing", file_path)


# Address Data
class Address:
    '''get Addresses
    file_name = 'Abonnenten Gebühren einzeilig.xlsx'
    writes address_data with abo_nr as key
    '''
    def load(self, file_name):
        # Load the workbook and select a sheet
        file_path = Path(
            settings.BASE_DIR) / 'billing' / 'fixtures' / file_name
        wb = load_workbook(file_path)
        ws = wb.active  # Or wb['SheetName']
        rows = [row for row in ws.iter_rows(values_only=True)]

        # Init
        address_data = {}

        # Read
        for row_nr, row in enumerate(rows):
            cells = row
            if cells[0] and (isinstance(cells[0], int) or isinstance(cells[0], float)):
                (
                    abo_nr, r_empf, persnr, namevorname, _, strasse, plz_ort, _,
                    tarif, periode, tarifbez, basis, ansatznr, ansatz, tage,
                    betrag, inklmwst, steuercodezähler, berechnungscodezähler,
                    steuercodegebühren, berechnungscodegebühren, gebührentext,
                     gebührenzusatztext
                ) = row
                if strasse:                   
                    strasse = strasse.replace('strase', 'strasse')
                    category = 'Allend' if 'Allend' in strasse else 'Gunzgen'
                else:
                    category = 'Gunzgen'
                address_data[abo_nr] = {
                    'plz': plz_ort.split(' ')[0],
                    'city': plz_ort.split(' ')[1],
                    'address': strasse,
                    'category': category
                }

        write(address_data, 'addresses.json')
        return address_data


# Product Data
class Product:
    '''
    many old products --> don't use
    file_name = 'Gebührentarife.xlsx'
    '''
    def load(self, file_name):
        # Load the workbook and select a sheet
        file_path = Path(
            settings.BASE_DIR) / 'billing' / 'fixtures' / file_name

        wb = load_workbook(file_path)
        ws = wb.active  # Or wb['SheetName']
        rows = [row for row in ws.iter_rows(values_only=True)]

        # Init
        self.products = []

        # Read
        for row_nr, row in enumerate(rows):
            cells = row
            if cells[1] == 'WA':
                (
                    period, category, _, name, _, _, _, _, _, account, _,
                    staffeltarif, _, price, *_
                ) = row
                self.products.append({
                    'period': period,
                    'category': category,
                    'name': name,
                    'price': price
                })

    def distinct(self):
        '''add:
            {'category': 'WA', 'name': 'Wasser Unterzähler', 'price': 1.1},
            {'category': 'WA', 'name': 'ARA Unterzähler', 'price': 2},
            {'category': 'WA', 'name': 'Mahngebühren', 'price': 20},
            {'category': 'WA', 'name': 'Verzugszinsen', 'price': None},
            {'category': 'WA', 'name': 'Mahngebühren WA', 'price': 20},
        '''
        distincts = []
        for product in [dict(x) for x in self.products]:
            product.pop('period')
            if product not in prod:
                distincts.append(product)

        return distincts


# Counter Data
class Counter:
    '''get Counter and product data from
        Abonnenten mit Zähler und Gebühren.xlsx
    '''
    file_name = 'Abonnenten mit Zähler und Gebühren.xlsx'
    city = 'Gunzgen'
    zip_code = 4617

    def __init__(self, category='WA'):
        self.category = category

    @staticmethod
    def get_company(name):
        if ' AG' in name:
            return name
        elif ' gmbh' in name.lower():
            return name
        elif 'genossenschaft' in name.lower():
            return name
        elif 'verband' in name.lower():
            return name
        elif 'verein' in name.lower():
            return name
        elif 'gemeinde' in name.lower():
            return name
        elif 'STWEG' in name.upper():
            return name

        return None

    @staticmethod
    def get_last_name(name):
        return name.split(' ', 1)[0]

    @classmethod
    def get_city(cls, address):
        return cls.city

    @classmethod
    def get_zip(cls, address):
        return cls.zip_code

    @classmethod
    def get_address(cls, address):
        if address:
            return address.replace(', Gunzgen', '')
        return None

    def load(self, address_data, file_name):
        # Load the workbook and select a sheet
        file_path = Path(
            settings.BASE_DIR) / 'billing' / 'fixtures' / file_name
        wb = load_workbook(file_path)
        ws = wb.active  # Or wb['SheetName']
        rows = [row for row in ws.iter_rows(values_only=True)]

        # Init
        address_categories = []   # -> AddressCategory
        addresses = []  # -> Address
        building = {}  # key: subscriber_number
        subscriber = {}  # key: subscriber_number
        subscription = {}
        counter = {}
        montage = {}
        product = {}
        measurements = []

        # Load
        for row_nr, row in enumerate(rows):
            cells = row
            if (cells[0] and isinstance(cells[0], str)
                    and cells[0].startswith('WA-')):
                # Subscription data
                (
                    period, subscriber_name, _, _, _, subscriber_number, *_
                ) = rows[row_nr]
                (
                    _, invoice_address, _, _, mut_c, *_
                ) = rows[row_nr + 1]
                (
                    _, _, _, _, _, start, _, _, _, _, _, building_address, *_
                ) = rows[row_nr + 2]
                (
                    _, invoice_receiver, _, _, _, exit, _, _, _, _, _,
                    building_category, *_
                ) = rows[row_nr + 3]

                address = {
                    'address': self.get_address(building_address),
                    'zip': self.get_zip(building_address),
                    'city': self.get_city(building_address),
                    'country': 'CHE'
                }

                if address not in addresses:
                    addresses.append(address)

                address_key = building_address
                building_key = subscriber_number

                building[subscriber_number] = {
                    'address': address_key,
                    'building_category': building_category,
                }

                company = self.get_company(subscriber_name)

                invoice_address = address_data.get(subscriber_number, None)
                if not invoice_address:
                    print("*problem", subscriber_number)

                subscriber[subscriber_number] = {
                    'company': company,
                    'last_name': (
                        None if company 
                        else self.get_last_name(subscriber_name)),
                    'alt_name': subscriber_name,
                    'address': {
                        'address': address_key,
                        'zip_code': self.zip_code,
                        'city': self.city,
                    },
                    'invoice_receiver': invoice_receiver,
                    'invoice_address': invoice_address,
                }

            elif isinstance(cells[0], int):
                if cells[0] > 100:
                    # Counter data
                    (
                        counter_nr, montage_nr, wohnungsbez, abl_code, _,
                        montage_date, _, tarif, bez, tage, fkt, anz_zw,
                        zw_alt_1, zw_alt_2, strg_z, add, zuge, zuga, folge
                    ) = row
                    counter[counter_nr] = {
                        'tarif': tarif,
                        'bez': bez,
                        'anz_zw': anz_zw,
                        'zw_alt_1': zw_alt_1,
                        'zw_alt_2': zw_alt_2,
                        'strg_z': strg_z,
                        'wohnungsbez': wohnungsbez,
                        'building': building_key,
                        'subscriber': subscriber_number
                    }
                    montage[counter_nr] = {
                        'date': montage_date,
                    }
                else:
                    # Pricing data
                    (
                        tarif, bez, p_text, _ , _ , basis, anr, ansatz, betrag,
                        tage, text, _ , zusatztext, _ , strgz, berz, strgg,
                        berg, folge
                    ) = row

                    # Product
                    product_key = f"{tarif}_{anr or '0'}"

                    description = '' if text or zusatztext else None
                    if text:
                        description = text
                    if zusatztext:
                        if description:
                            description += ', ' + zusatztext
                        else:
                            description = zusatztext

                    product[product_key] = {
                        'category': self.category,
                        'tarif': tarif,
                        'bez': bez,
                        'anr': anr,
                        'name': p_text,
                        'description': description,
                        'price': ansatz,
                        'strgz': strgz,
                        'berz': berz,
                    }

                    # measurements
                    if tarif in [14, 20]:
                        measurements.append({
                            'counter': counter_nr,
                            'period': period,
                            'value': (zw_alt_1 if zw_alt_1 else 0) +
                                (basis if basis else 0),
                            'consumption': basis
                        })

                    # Subscription
                    if subscriber_number not in subscription:
                        subscription[subscriber_number] = {
                            'subscriber': subscriber_number,
                            'building': subscriber_number,
                            'products': [],
                            'start': start,
                            'exit': exit,
                        }                       
                    subscription[subscriber_number]['products'].append(
                        product_key)

        # Data
        data = dict(
            address_categories=address_categories,
            addresses=addresses,
            building=building,
            subscriber=subscriber,
            subscription=subscription,
            counter=counter,
            montage=montage,
            product=product,
            measurements=measurements
        )

        write(data, 'counter.json')
