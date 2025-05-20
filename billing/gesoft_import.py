'''
billing/gesoft_import.py
'''
from enum import Enum
import json
import logging
import re
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from pathlib import Path

from django.conf import settings
from django.utils import timezone

from accounting.models import ArticleCategory, Article
from asset.models import (
    DEVICE_STATUS, Unit, AssetCategory, Device, EventLog)
from core.models import (
     Tenant, AddressMunicipal, Area, Address, PersonCategory,
     Person, PersonAddress
)
from scerp.mixins import parse_gesoft_to_datetime
from .models import (
    ARTICLE_NR_POSTFIX_DAY, Period, Route, Measurement, Subscription,
    SubscriptionArticle, SubscriptionArchive
)

logger = logging.getLogger(__name__)


# Sync
ARTICLE_IS_ENABLED_SYNC = True  # sync articles
PERSON_IS_ENABLED_SYNC = True


# Address default
CITY = 'Gunzgen'
ZIP = 4617
ALLMEND_EGIDS = [
    502181189,
    9041136,
    502181223,
    502181108,
    192039736,
    502181109,
    502180980,
    2125744
]

class AREA(Enum):
    GUNZGEN = {'code': 'gunzgen', 'name': 'Gunzgen'}
    ALLMEND = {'code': 'allmend', 'name': 'Allmend'}


ADDITIONAL_EGIDS = [
    502181189,
    2125744,
    502181035,
    502181163,
    9023745,
    192039728,
    502181164,
    502180979,
    9041136,
    502180980,
    502181035,
    502181163,
    9023745,
    9023745,
    192039728,
    502181164,
    502180979,
    9041136,
    502180980
]


# Needed for area assignment
BUILDING_MAP = {
    # AboNr: Address in 4617, or EGID Number',
    44: 'Industriestrasse 7',
    556: 'Markstrasse 12',
    409: 'Raststätte Gunzgen Nord 2',
    602: 'Markstrasse 12',
    624: 'Bornstrasse 14',
    488: 'Banackerstrasse 25',
    457: 'Niederhofweg 4',
    65: 'Niederhofweg 1',
    64: 'Industriestrasse 2',
    63: 'Industriestrasse 4',
    62: 'Niederhofweg 5',
    61.1: 'Niederhofweg 13',
    61: 'Niederhofweg 11',
    206: 'Alte Poststrasse 4',
    204: 'Schulstrasse 2',
    203: 'Allmendstrasse 8',
    202: 'Allmendstrasse 2',
    201: 'Schulstrasse 4',
    200: 'Ghölstrasse 20',
    288: 'Rotsangelstrasse 13',
    533: 'Ghölstrasse 18',
    572: 'Industriestrasse 19',
    142: 'Allmendstrasse 31',
    628: 'Niderfeld 2',
    42: 'Industriestrasse 23',
    458: 'Härkingerstrasse 1l',
    408: 'Härkingerstrasse 1i',
    407.1: 'Härkingerstrasse 1m',
    407: 'Härkingerstrasse 1g',
    313: 'Niderfeld 9',
    86.1: 'Römerweg 20',
    597: 'Mittelgäustrasse 81',
    557: 'Markstrasse 16',
    212: 'Hanselmattweg 4',
    81: 'Oberfeldweg 24',
    404: 'Raststätte Gunzgen Süd 3',
    23: 'Lipsmattweg 3',
    427: 'Eichliban 1',
    403: 'Raststätte Gunzgen Nord 2.1',
    401: 'Allmend 29',
    45: 'Industriestrasse 11',
    127: 'Römerweg Ost 20',
    431: 'Banackerstrasse 28a',
    438: 'Banackerstrasse 26a',
    351: 'Mittelgäustrasse 52',
    167: 'Banackerstrasse 24a',
    280: 'Mittelgäustrasse 37',
    626.2: 'Raststätte Gunzgen Süd 1a',
    626.1: 'Raststätte Gunzgen Süd 1',
    343: 'Mittelgäustrasse 49',
    417: 'Klärstrasse 12',
    205: 'Klärstrasse 12',
    288: 502272780,
    404: 9041136,
    626.1: 2125744,
    626.2: 502180980
}

# Wrong Counters
COUNTER_REPLACE = {
    # old --> new
    (23529877, 1276): (23529878, 1276)
}

COMPANIES = [
    'GmbH',
    'AG',
    ' SA',
    'Einwohnergemeinde',
    'Eigentümergemeinschaft',
    'Erbengemeinschaft',
    'Genossenschaft',
    'Kirchgemeinde',
    'Zweckverband',
    'Bürgergemeinde',
    'Immobilien',
    'STOWE',
    'STWE'
    'STWEG'
]


ARTICLE_MAPPING = {
    # Tarif, Ansatz
    ('WATER_COLD', 1.1): {
            'number': 'A-W-011',
            'name': {'de': 'Gebühr Wasser'},
            'category': 'water_cold',
            'unit': 'volume'
        },
    ('WATER_HOT', 1.1): {
            'number': 'A-WH-021',
            'name': {'de': 'Gebühr Warmwasser'},
            'category': 'water_hot',
            'unit': 'volume'
        },
    (5, 22.5): {
            'number': 'A-WA-051',
            'name': {'de': 'Zählermiete Wasser'},
            'category': 'water_cold',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (5, 35): {
            'number': 'A-WA-052',
            'name': {'de': 'Zählermiete Wasser Industrie'},
            'category': 'water_cold',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (6, 1953): {
            'number': 'A-WA-061',
            'name': {'de': 'Sprinkleranlage'},
            'category': 'water_cold',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (12, None): {
            'number': 'A-WW-120',
            'name': {'de': 'Grundgebühr Abwasser Wohnung - befreit'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (12, 60): {
            'number': 'A-WW-121',
            'name': {'de': 'Grundgebühr Abwasser Wohnung'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 60): {
            'number': 'A-WW-131',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T1'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 180): {
            'number': 'A-WW-132',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T2'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 375): {
            'number': 'A-WW-133',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T3'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 555): {
            'number': 'A-WW-134',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T4'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 825): {
            'number': 'A-WW-135',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T5'},
            'unit': 'period',
            'category': 'water_waste'
        },
    (13, 840): {
            'number': 'A-WW-136',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T6'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 880): {
            'number': 'A-WW-137',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T7'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 1140): {
            'number': 'A-WW-138',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T8'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (13, 2215): {
            'number': 'A-WW-139',
            'name': {'de': 'Grundgebühr Abwasser Industrie - T9'},
            'category': 'water_waste',
            'unit': 'period',
            'notes': 'Einheit: Periode oder Tage'
        },
    (14, 1.4): {
            'number': 'A-WW-141',
            'name': {'de': 'Gebühr Abwasser'},
            'category': 'water_waste',
            'unit': 'volume'
        },
    (20, 2.45): {
            'number': 'A-WW-201',
            'name': {'de': 'Gebühr Abwasser T1.75x'},
            'category': 'water_waste',
            'unit': 'volume'
        }
}

# helpers
def is_numeric_string(s):
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


class Import:

    def __init__(self, tenant_id):
        '''
            datetime_default,e g. '2024-03-31'
        '''
        # From tenant
        self.tenant = Tenant.objects.get(id=tenant_id)
        self.created_by = self.tenant.created_by
        self.person_category = PersonCategory.objects.get(
            tenant=self.tenant, code='subscriber')

    @staticmethod
    def clean_address(address):
        if address:
            address = address.replace('Gunzgen', '')
            address = address.replace('strase', 'strasse')
            address = address.replace('str.', 'strasse')
            address = address.strip()
            if address[-1] == ',':
                address = address[:-1].strip()

        return address

    @staticmethod
    def clean_cell(value):
        if value and isinstance(value, str):
            return value.strip()
        return value


class ImportAddress(Import):

    def load(self, file_name):
        '''get Addresses
        we use this to get the invoicing addresses

        file_name = 'Abonnenten Gebühren einzeilig.xlsx'
        writes address_data with abo_nr as key
        '''
        file_path = Path(
            settings.BASE_DIR) / 'billing' / 'fixtures' / file_name
        wb = load_workbook(file_path)
        ws = wb.active  # Or wb['SheetName']
        rows = [row for row in ws.iter_rows(values_only=True)]

        # Read
        address_data = {}  # key name
        for row_nr, row in enumerate(rows, start=1):
            cells = row
            if cells[0] and isinstance(cells[0], (int, float)):
                # Get data
                (abo_nr, _, _, namevorname, _, strasse, plz_ort, *_) = row
                namevorname = namevorname.strip()

                # Make address
                if not strasse:
                    # e.g. Autobahnraststätte Gunzgen Nord AG
                    strasse = namevorname.strip()

                # Save address
                data = {
                    'name': namevorname,
                    'zip': plz_ort.split(' ')[0],
                    'city': plz_ort.split(' ')[1],
                    'address': self.clean_address(strasse)
                }

                if (namevorname == 'Bürgergemeinde Gunzgen'
                        and data['zip'] == '4616'):
                     logging.warning(
                        f"{row_nr} {namevorname} {data['zip']} ignoring")
                     continue

                if namevorname in address_data:
                    if address_data[namevorname] != data:
                        logging.warning(
                            f"{row_nr} {namevorname} data['zip'] "
                            "has no unique address")

                address_data[namevorname] = data

        return address_data


class AreaAssignment(ImportAddress):

    def __init__(self, tenant_id):
        super().__init__(tenant_id)

    def assign(self):
        # Area
        area_obj = {}
        for item in AREA:
            area = item.value
            obj, _created = Area.objects.get_or_create(
                tenant=self.tenant,
                code=area['code'],
                defaults={
                    'name': area['name'],
                    'created_by': self.created_by
                }
            )
            area_obj[obj.code] = obj

        # Assign Areas
        count = 0
        for address in AddressMunicipal.objects.filter(tenant=self.tenant):
            if (address.stn_label == AREA.ALLMEND.value['name']
                    or address.bdg_egid in ALLMEND_EGIDS):
                address.area = area_obj[AREA.ALLMEND.value['code']]
            else:
                address.area = area_obj[AREA.GUNZGEN.value['code']]
            address.save()
            count += 1

        logging.info(f"Updated {count} addresses")


class ImportData(ImportAddress):

    def __init__(self, tenant_id, route_id, datetime_default):
        '''
            datetime_default,e g. '2024-03-31'
        '''
        super().__init__(tenant_id)

        # From route
        self.route = Route.objects.get(
            tenant=self.tenant, id=route_id)
        self.period = self.route.period
        self.asset_categories = self.route.asset_categories.all()

        # Time
        self.datetime = timezone.make_aware(
            datetime.strptime(datetime_default, "%Y-%m-%d"))
        self.date = self.datetime.date()

    @staticmethod
    def convert_to_date(date_string):
        ''' e.g. date_string = "22.04.2009"
        '''
        return datetime.strptime(date_string, '%d.%m.%Y').date()

    @staticmethod
    def parse_subscriber_name(subscriber_name):
        """
        Parses the subscriber name into a person and partner dictionary.
        Returns:
            person: Dictionary containing company, title, last_name, and first_name of the main person.
            partner: Dictionary containing last_name and first_name of the partner (if available).
        """
        subscriber_name = subscriber_name.strip()

        person = {
            'subscriber_name': subscriber_name,
            'company': None,
            'title': None,
            'last_name': None,
            'first_name': None,
            'partner_title': None,
            'partner_last_name': None,
            'partner_first_name': None,
        }

        # Clean
        subscriber_name = subscriber_name.replace('  ', ' ')

        # Condition 1: If name contains "GmbH" or "AG" (exactly)
        for keyword in COMPANIES:
            if keyword in subscriber_name:
                person['company'] = subscriber_name.strip()
                return person

        # Condition 2: Matches "last_name first_name und first_name_partner" or "last_name first_name + first_name_partner"
        pattern = (
            r"^(?P<last_name>[^\s]+)\s(?P<first_name>[^\s]+)\s(?:und|\+ ?)"
            r"(?P<partner_first_name>.+)$"
        )
        match = re.match(pattern, subscriber_name)
        if match:
            person.update({
                'last_name': match.group("last_name"),
                'first_name': match.group("first_name"),
                'partner_last_name': match.group("last_name"),
                'partner_first_name': match.group("partner_first_name").strip(),
            })
            return person

        # Condition 3: Matches "last_name first_name" (single person)
        pattern_single = r"^(?P<last_name>[^\s]+)\s(?P<first_name>.+)$"
        match_single = re.match(pattern_single, subscriber_name)
        if match_single:
            person.update({
                'last_name': match_single.group("last_name"),
                'first_name': match_single.group("first_name").strip(),
            })
            return person

        # Condition 4: If no pattern matches, treat the entire string as the last name
        person['last_name'] = subscriber_name.strip()

        return person

    # Models

    # Core
    def add_address(self, data):
        data['created_by'] = self.created_by
        obj, created = Address.objects.get_or_create(
            tenant=self.tenant,
            zip=data.pop('zip'),
            city=data.pop('city'),
            address=data.pop('address'),
            defaults=data
        )
        return obj, created

    def add_person(self, data):
        if data['company']:
            # find
            query_set = Person.objects.filter(
                tenant=self.tenant,
                category=self.person_category,
                company=data['company']
            )
        else:
            #  We hope that nobody has the same first and last name
            query_set = Person.objects.filter(
                tenant=self.tenant,
                category=self.person_category,
                last_name=data['last_name'],
                first_name=data['first_name'],
            )

        if query_set:
            # update
            query_set.update(
                company=data['company'],
                last_name=data['last_name'],
                first_name=data['first_name'],
                notes=data['notes']
            )
            person = query_set.first()
            created = False
        else:
            person = Person.objects.create(
                tenant=self.tenant,
                company=data['company'],
                last_name=data['last_name'],
                first_name=data['first_name'],
                notes=data['notes'],
                category=self.person_category,
                is_customer=True,
                is_enabled_sync=PERSON_IS_ENABLED_SYNC,
                sync_to_accounting=PERSON_IS_ENABLED_SYNC,
                created_by=self.created_by
            )
            created = True

        return person, created

    def add_person_address(
            self, person, address, address_type, additional_information=None):
        obj, created = PersonAddress.objects.get_or_create(
            tenant=self.tenant,
            address=address,
            type=address_type,
            person=person,
            defaults=dict(
                additional_information=additional_information,
                created_by=self.created_by
            )
        )

        return obj, created

    # Asset
    def add_device(self, data):
        device = Device.objects.filter(
            tenant=self.tenant,
            code=data.pop('code')
        ).first()
        if device:
            return device

        raise ValueError(f"counter {obj.code} does not exist.")

    def add_event(self, device, dt, status, data):
        data['created_by'] = self.created_by
        obj, created = EventLog.objects.get_or_create(
            tenant=self.tenant,
            device=device,
            datetime=dt,
            status=status,
            defaults=data
        )
        return obj, created

    # Accounting
    def add_article(self, tarif, price):
        # Init
        article = ARTICLE_MAPPING[tarif, price]
        category = ArticleCategory.objects.get(
            tenant=self.tenant, code=article['category'])

        # Save data
        obj, created = Article.objects.get_or_create(
            tenant=self.tenant,
            nr=article['number'],
            defaults={
                'name': article['name'],
                'category': category,
                'sales_price': price or 0,
                'unit': Unit.objects.get(
                    tenant=self.tenant, code=article['unit']),
                'created_by': self.created_by,
                'is_enabled_sync': ARTICLE_IS_ENABLED_SYNC,
                'sync_to_accounting': ARTICLE_IS_ENABLED_SYNC
            }
        )
        return obj, created

    # Billing
    def add_measurement(self, counter, route, datetime, data):
        data['created_by'] = self.created_by
        obj, created = Measurement.objects.get_or_create(
            tenant=self.tenant,
            counter=counter,
            route=route,
            datetime=datetime,
            defaults=data
        )
        return obj, created

    def add_subscription(self, subscriber_number, data):
        obj, created = Subscription.objects.update_or_create(
            tenant=self.tenant,
            created_by=self.created_by,
            subscriber_number=subscriber_number,
            defaults=data
        )
        return obj, created

    def add_subscription_article(self, subscription, article):
        subscription.articles.add(article)

    def load_block_intro(self, row_nr, rows, address_data):
        ''' load first block, 4 lines
            returns:
                subscriber_number
                building
                subscription

        '''
        (
            period, subscriber_name, _, _, _, subscriber_number, *_
        ) = rows[row_nr]
        (
            _, subscriber_address, _, _, mut_c, *_
        ) = rows[row_nr + 1]
        (
            _, _, _, _, _, start, _, _, _, _, _, building_address, *_
        ) = rows[row_nr + 2]
        (
            _, invoice_name, _, _, _, exit, _, _, _, _, _,
            building_category, *_
        ) = rows[row_nr + 3]

        # Clean
        subscriber_name = self.clean_cell(subscriber_name)
        subscriber_address_c = self.clean_address(subscriber_address)
        invoice_name = self.clean_cell(invoice_name)
        building_address_excel = building_address

        if subscriber_number in BUILDING_MAP:
            building_address_c = BUILDING_MAP[subscriber_number]
        elif building_address:
            building_address, *building_category = (
                building_address.split(',', 1))   # sometime in one cell
            if not building_address:
                building_address, *building_category = (
                    building_address.split('    ', 1))   # sometime in one cell
            if not building_address:
                building_address, *building_category = (
                    building_address.split('    ', 1))   # sometime in one cell
            building_address_c = self.clean_address(building_address)

        if not building_category:
            building_category = self.clean_cell(building_category)

        # logger.info(f"Reading abo_nr {subscriber_number}")

        # Building Address
        try:
            stn_label, adr_number = building_address_c.rsplit(' ', 1)
        except:
            stn_label, adr_number = None, None

        if isinstance(building_address_c, int):
            addr_building = AddressMunicipal.objects.filter(
                tenant=self.tenant, bdg_egid=building_address_c
            ).first()
        elif stn_label:
            addr_building = AddressMunicipal.objects.filter(
                tenant=self.tenant, zip=ZIP,
                stn_label=stn_label, adr_number=adr_number).first()
        else:
            addr_building = None

        # Subscriber
        person = self.parse_subscriber_name(subscriber_name)

        data = {
            'company': person['company'],
            'last_name': person['last_name'],
            'first_name': person['first_name'],

            # we use notes for storing old number
            'notes': (
                f"abo_nr: {subscriber_number}, "
                f"subscriber_name: {subscriber_name}")
        }
        subscriber, _created = self.add_person(data)

        # Partner
        if person['partner_last_name']:
            data = {
                'company': None,
                'last_name': person['partner_last_name'],
                'first_name': person['partner_first_name'],

                # we use notes for storing old number
                'notes': (
                    f"abo_nr: {subscriber_number}, "
                    f"subscriber_name: {subscriber_name}")
            }
            partner, _created = self.add_person(data)
        else:
            partner = None

        # Get subscriber address, we look it up from address_data
        lookup = address_data.get(subscriber_name)
        if lookup:
            subscriber_address_data = dict(
                zip=lookup['zip'],
                city=lookup['city'],
                address=lookup['address']
            )
        else:
            raise ValueError(f"No address found for {subscriber_name}")

        # Add subscriber address
        address, _created = self.add_address(subscriber_address_data)
        subscriber_address, _created = self.add_person_address(
            subscriber, address, PersonAddress.TYPE.MAIN)
        if partner:
            partner_address, _created = self.add_person_address(
                partner, address, PersonAddress.TYPE.MAIN)

        # Invoice address
        # Check if invoice_name == subscriber_name
        invoice_note = None
        if (subscriber_address_c != building_address_c):
            lookup = address_data.get(invoice_name)
            if lookup:
                invoice_address_data = dict(
                    zip=lookup['zip'],
                    city=lookup['city'],
                    address=f"{lookup['address']}"
                )
                invoice_note = 'check invoice address, '
            else:
                logger.warning(f"No address found for {invoice_name}'not")
                invoice_address_data = dict(
                    zip='0000',
                    city='???',
                    address='???'
                )
                invoice_note = 'invoice address not found, '

            # add invoice address
            address, _created = self.add_address(invoice_address_data)
            invoice_address, _created = self.add_person_address(
                subscriber, address, PersonAddress.TYPE.INVOICE,
                invoice_name)

        # Subscription
        start_date = self.convert_to_date(start)
        if exit:
            end_date = self.convert_to_date(exit)
        else:
            end_date = None

        data = {
            'subscriber': subscriber,
            'partner': partner,
            'address': addr_building,
            'start': start_date,
            'end': end_date,
            'notes': ''
        }
        if addr_building:
            data['notes'] = building_address_excel
        else:
            data['notes'] = f"{building_address_excel}: building not found"
            logger.info(f"{building_address_excel} building not found")
        if invoice_note:
            if data['notes']:
                data['notes'] += ', '
            data['notes'] += invoice_note
        if building_category:
            data['notes'] += f'building: {building_category}'

        subscription, _created = self.add_subscription(
            subscriber_number, data)

        return subscriber_number, addr_building, subscription

    def load_block_counter(self, row, addr_building, subscription):
        '''
        returns:
            water_tarif
            counter
            measurement
        '''
        (
            counter_nr, montage_nr, wohnungsbez, abl_code, _,
            montage_date, _, tarif, bez, tage, fkt, anz_zw,
            zw_alt_1, zw_alt_2, strg_z, add, zuge, zuga, folge
        ) = row

        # Add Counter
        dt = parse_gesoft_to_datetime(montage_date)
        data = {
            'code': counter_nr,
            'number': counter_nr,
            'date_added': dt.date(),
        }

        if (counter_nr, montage_nr) in COUNTER_REPLACE:
            # Counter replace
            counter_nr, montage_nr = COUNTER_REPLACE[(counter_nr, montage_nr)]
            logger.info(f"counter replacement {counter_nr}")

        # Check device
        device = self.add_device(data)

        # Add to subscription
        subscription.counters.add(device)

        # Add Montage
        data = {
            'address': addr_building,
            'notes': f'Mont-Nr. {montage_nr}'
        }
        event, _created = self.add_event(
            device, dt, DEVICE_STATUS.MOUNTED, data)

        # Add Measurements
        data = {
            'value_previous': zw_alt_1 or zw_alt_2,
            'address': addr_building,  # efficieny
            'period': self.period,  # efficieny
            'subscription': subscription,  # efficieny
        }
        measurement, _created = self.add_measurement(
            device, self.route, self.datetime, data)

        return tarif, device, measurement

    def load_block_pricing(
            self, row, subscription, measurement, consumption_only=False):
        '''use consumption_only if values should be not updated and it is
            just for statictics; these records needs to be manually updated
        '''
        (
            tarif, bez, p_text, _ , _ , basis, anr, ansatz, betrag,
            tage, text, _ , zusatztext, _ , strgz, berz, strgg,
            berg, folge
        ) = row

        # Measurement
        if measurement and tarif in [14, 20]:
            # Update consumption
            measurement.consumption = basis or 0

            # Update value
            if not consumption_only:
                # Update value
                value_previous = measurement.value_previous or 0
                measurement.value = value_previous + measurement.consumption

                # Update value_min, value_max
                self.value_min = (
                    measurement.consumption * self.route.confidence_min)
                self.value_max = (
                    measurement.consumption * self.route.confidence_max)

            measurement.save()

        # Article
        article, _created = self.add_article(tarif, ansatz)

        # Add article
        self.add_subscription_article(subscription, article)

    def load(self, file_name, address_data):
        '''
        Note: we do not know the inhabitant if not identical with subscriber
        '''
        # Load the workbook and select a sheet
        file_path = Path(
            settings.BASE_DIR) / 'billing' / 'fixtures' / file_name
        wb = load_workbook(file_path)
        ws = wb.active  # Or wb['SheetName']
        rows = [row for row in ws.iter_rows(values_only=True)]

        # Make water articles if not done
        article_water_cold, _created = self.add_article('WATER_COLD', 1.1)
        article_water_hot, _created = self.add_article('WATER_HOT', 1.1)

        # Pre-Load
        for row_nr, row in enumerate(rows):
            first_cell = row[0]
            if (first_cell and isinstance(first_cell, str)
                    and first_cell.startswith('WA-')):
                # Load intro
                result = self.load_block_intro(row_nr, rows, address_data)

        # Load
        for row_nr, row in enumerate(rows):
            first_cell = row[0]
            if (first_cell and isinstance(first_cell, str)
                    and first_cell.startswith('WA-')):
                # Load intro
                result = self.load_block_intro(row_nr, rows, address_data)
                subscriber_number, addr_building, subscription = result

                # check if
                if not addr_building:
                    logger.warning(
                        f"No building address found for {subscriber_number}. "
                        "Cannot continue processing."
                    )

                measurements = []
                tarif_water = None
            elif isinstance(first_cell, (int, float)) and addr_building:
                if first_cell > 100:
                    # Load counters
                    result = self.load_block_counter(
                        row, addr_building, subscription)
                    tarif_water, counter, measurement = result

                    # Add water article
                    if counter.category.code == '9-0:1.0.0':
                        article = article_water_hot
                    else:
                        article = article_water_cold
                    self.add_subscription_article(subscription, article)

                    # Add measurement
                    measurements.append(measurement)

                else:
                    # Load pricing
                    if measurements:
                        # we take the first

                        # If there are more than one counter we just take the
                        # first one and put the total consumption in the first
                        # counter; manually update later!!!
                        update_measurement = measurements[0]
                        consumption_only = len(measurements) > 1

                        # load consumption
                        self.load_block_pricing(
                            row, subscription, update_measurement,
                            consumption_only)
                    else:
                        msg = f"Row nr {row_nr}: no measurement created for {row}"
                        logger.warning(msg)


class ArticleCopy(Import):

    def make_daily(self):
        '''
        look for period articles and make day articles
        '''
        # Init
        articles = Article.objects.filter(
            tenant=self.tenant, unit__code='period')
        article_nr_copies = [
            article.nr
            for article in Article.objects.filter(
                tenant=self.tenant, nr__endswith=ARTICLE_NR_POSTFIX_DAY)
        ]
        unit = Unit.objects.filter(
            tenant=self.tenant, code='day').first()

        # Copy
        count = 0
        for article in articles:
            # assign and check nr
            article.nr += ARTICLE_NR_POSTFIX_DAY
            article.name = {
                lang: name + ARTICLE_NR_POSTFIX_DAY
                for lang, name in article.name.items()
            }

            if article.nr not in article_nr_copies:
                # make a copy
                article.pk = None
                article.c_id = None
                article.sync_to_accounting = True
                article.sales_price = float(article.sales_price / 180)
                article.unit = unit
                article.save()
                count += 1
                logger.info(f"saving {article.nr}")

        return count

    def rename_daily(self):
        '''
        rename day articles
        '''
        # Init
        articles = Article.objects.filter(
            tenant=self.tenant, unit__code='day')

        # Copy
        count = 0
        for article in articles:
            # new name
            article.name = {
                lang: name + ARTICLE_NR_POSTFIX_DAY
                for lang, name in article.name.items()
            }

            # save
            article.save()
            count += 1
            logger.info(f"saving {article.nr}")

        return count


class ImportArchive(Import):
    '''class for archiving, load after actual Subscriptions have been loaded
    '''
    def load(self, file_name):
        '''get Archive data

        file_name = 'Abonnenten Archiv Gebühren einzeilig.xlsx'
        '''
        file_path = Path(
            settings.BASE_DIR) / 'billing' / 'fixtures' / file_name
        wb = load_workbook(file_path)
        ws = wb.active  # Or wb['SheetName']
        category = ArticleCategory.objects.get(
            tenant=self.tenant, code='archive')

        # Read
        count = 0
        for row in ws.iter_rows(values_only=True):
            cells = row
            if cells[0] and isinstance(cells[0], (int, float)):
                # Get data
                (abo_nr, r__empf, pers_nr, name_vorname, _, strasse, plz_ort,
                 _, tarif, periode, tarif_bez, basis, ansatz_nr, ansatz, tage,
                 betrag, inkl_mw_st, steuercode_zaehler,
                 berechnungscode_zaehler, steuercode_gebuehren,
                 berechnungscode_gebuehren, gebuehrentext,
                 gebuehren_zusatztext, *_) = row
                """
                # Subscriber
                subscription = Subscription.objects.filter(
                    tenant=self.tenant, subscriber_number=abo_nr).first()
                if not subscriber:
                    # Person
                    person = Person.objects.create(
                        tenant=self.tenant,
                        created_by=self.created_by,
                        category=self.person_category,
                        last_name=f'*archive {abo_nr}',
                        notes = f'''
                            Subscriber: {name_vorname}
                            Addresse: {strasse}, {zip} {city}
                            '''
                    )

                    # Subscription
                    subscription = Subscription.objects.create(
                        tenant=self.tenant,
                        created_by=self.created_by,
                        subscriber_number=abo_nr,
                        subscriber=person
                    )

                # Article
                obj, created = Article.objects.get_or_create(
                    tenant=self.tenant,
                    nr='A-W-A99',
                    defaults={
                        'name': f"{tarif}_{ansatz_nr}",
                        'category': category,
                        'sales_price': betrag or 0,
                        'created_by': self.created_by,
                        'is_enabled_sync': True,
                        'sync_to_accounting': True
                    }
                )

                if steuercode_gebuehren != 'A':
                    continue  # no measurement
                """
                # Period
                period_str = str(periode)
                if len(period_str) < 4:
                    period_str = '0' + period_str

                year = 2000 + int(period_str[:2])
                month = 4 if period_str[-1] == '1' else 10
                day = 1
                start = date(year, month, day)

                if tage:
                    end = start + timedelta(days=tage)
                else:
                    year = year if period_str[-1] == '1' else year + 1
                    month = 9 if period_str[-1] == '1' else 3
                    day = 30 if period_str[-1] == '1' else 31
                    try:
                        end = date(year, month, day)
                    except ValueError as e:
                        raise Exception(f"Invalid date created: {e}")

                period_obj, created = Period.objects.get_or_create(
                    tenant=self.tenant,
                    code=period_str,
                    defaults=dict(
                        name=f"Wasser {period_str}",
                        start=start,
                        end=end,
                        created_by=self.created_by
                    ))
                if created:
                    logging.info(f"Created {period_obj}")

                # Route
                route_obj, created = Route.objects.get_or_create(
                    tenant=self.tenant,
                    period=period_obj,
                    defaults=dict(
                        name=f"Wasser, Route {period_str}",
                        start=start,
                        end=end,
                        status=Route.STATUS.INVOICES_GENERATED,
                        created_by=self.created_by
                    ))
                '''
                # Measurement
                obj, created = Measurement.objects.get_or_create(
                    tenant=self.tenant,
                    route=route_obj,
                    subscription=subscription,
                    defaults={
                        'datetime': end,
                        'consumption'
                        'street_name': strasse,
                        'zip_city': plz_ort,
                        'tarif_name': tarif_bez,
                        'consumption': basis,
                        'amount': betrag,
                        'amount_gross': inkl_mw_st,
                        'created_by': self.tenant.created_by
                    }
                )
                '''
                if created:
                    count += 1

        logger.info(f"archived {count} positions")
        return count


def fix_zero_problem(json_filename, excel_file_name, tenant_id):
    ''' use to enter old  '''
    # load excel
    file_path = Path(
        settings.BASE_DIR) / 'billing' / 'fixtures' / excel_file_name
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active  # Or wb['SheetName']

    # Read
    analyse_list = [
        ['abo_nr', 'subscriber_excel', 'counter_nr', 'existing',
         'consumption_old', 'consumption_new', 'counters', 'products']
    ]
    abo_nr, subscriber_excel = None, None
    for row_nr, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header
        cell = row[0]
        value = cell.value
        number_format = cell.number_format

        # Reconstruct visible value
        if isinstance(value, str):
            visible_value = value
            # Get abo
            if value == 'WA-2402':
                abo_nr = row[5].value
                subscriber_excel = row[1].value
                continue

        elif isinstance(value, int):
            # If number format includes zero-padding, reconstruct that
            if '0' in number_format:
                zero_count = number_format.count('0')
                visible_value = str(value).zfill(zero_count)
            else:
                visible_value = str(value)

        elif isinstance(value, float):
            visible_value = str(value)  # Keep full float value like 18635332.1

        else:
            continue  # Skip None or unsupported types

        # Skip if no number
        if not is_numeric_string(visible_value):
            continue

        # Add products
        if float(visible_value) < 100:
            products = analyse_list[-1][-1]
            analyse_list[-1][-1] = products + 1
            continue

        # Get device
        counter_nr = visible_value
        devices = Device.objects.filter(
            tenant__id=tenant_id,
            code=counter_nr
        )
        if devices.count() > 1:
            analyse_list.append([
                abo_nr, subscriber_excel, counter_nr, 'multiple',
                None, None, 0, 0])
            continue
        elif not devices:
            analyse_list.append([
                abo_nr, subscriber_excel, counter_nr, None,
                None, None, 0, 0])
            continue
        device = devices.first()

        # Old Measurement
        measurements = Measurement.objects.filter(
            tenant__id=tenant_id,
            counter=device,
            datetime__lte=datetime(2024, 10, 1)
        )
        if measurements.first():
            consumption_old = measurements.first().consumption
        else:
            consumption_old = None

        # New Measurement
        measurements = Measurement.objects.filter(
            tenant__id=tenant_id,
            counter=device,
            datetime__gte=datetime(2025, 1, 1)
        )
        if measurements.first():
            consumption_new = measurements.first().consumption
        else:
            consumption_new = None

        analyse_list.append(
            [abo_nr, subscriber_excel, counter_nr, devices.exists(),
             consumption_old, consumption_new, 0, 0])

    # Calc counter_number
    for row in analyse_list[1:]:
        abo_nr = row[0]
        nr_of_counters = len([x for x in analyse_list if x[0] == abo_nr])
        row[-2] = nr_of_counters

    # Create a new workbook and get the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Report"

    # Add your data rows
    print("*analyse_list", analyse_list[:10])
    for row in analyse_list:
        ws.append(list(row))

    # Save to file
    file_name = 'analysis_report.xlsx'
    file_path = Path(
            settings.BASE_DIR) / 'billing' / 'fixtures' / file_name
    wb.save(file_path)

    logging.info(f"{file_name} created")


def adjust_articles(tenant_id):
    subscriptions = Subscription.objects.filter(tenant__id=tenant_id)
    for subscription in subscriptions.all():
        for article in subscription.articles.all():
            unit_code = article.unit.code
            quantity = 1 if unit_code in ['day', 'period'] else None                               
            obj, _created = SubscriptionArticle.objects.get_or_create(                
                tenant=subscription.tenant,
                subscription=subscription,
                article=article,
                defaults=dict(
                    quantity=quantity,
                    created_by=subscription.created_by
                )
            )            
            if _created:
                logging.info(f"{obj} created")
