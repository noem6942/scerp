'''
billing/calc.py
'''
from datetime import datetime
from decimal import Decimal
import io
import json
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, numbers

from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Sum, Min, Max, Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext as _

from accounting.models import Article, OutgoingOrder, OutgoingItem
from asset.models import Device
from core.models import (
    TenantSetup, Area, AddressMunicipal, PersonAddress, Attachment)
from scerp.admin import ExportExcel
from scerp.mixins import SafeDict, format_date, primary_language
from .models import (
    ARTICLE_NR_POSTFIX_DAY, Period, Route, Measurement, Subscription,
    SubscriptionArticle
)

logger = logging.getLogger(__name__)

# Defaults
CUSTOM_NUMBER_FORMAT = "#'##0"
CUSTOM_NUMBER_FORMAT_SMALL = "# ##0"

# GFT Template
class METER:
    TEMPLATE = {
        'billing_mde': {
            'route': {
                'name': None,
                'user': None
            },
            'meter': []
        }
    }


# Helpers
def calculate_growth(consumption_now, consumption_previous):
    """
    Calculates the percentage growth between two consumption values.
    Returns None if the previous consumption is zero (to avoid division by zero).
    """
    if consumption_previous == 0:
        return None  # Avoid division by zero
    return (
        (consumption_now - consumption_previous) / consumption_previous
    ) * 100


def convert_datetime_to_date(dt):
    return dt.strftime('%Y-%m-%d')


def convert_str_to_datetime(dt_string, date_format='%Y-%m-%d'):
    # is only date
    try:
        naive_datetime = datetime.strptime(dt_string, date_format)
    except:
        return None

    # Convert the naive datetime to an aware datetime
    return timezone.make_aware(naive_datetime)


def extract_datetime_from_route_filename(filename):
    ''' e.g. filename = "route0120250325_2025-03-25_16-28-49.json"
    '''
    try:
        # Extract the datetime string after the first underscore
        datetime_part = (
            filename.split('_')[1] + '_'
            + filename.split('_')[2].split('.')[0])

        # Parse the datetime string into a naive datetime object
        naive_datetime = datetime.strptime(datetime_part, '%Y-%m-%d_%H-%M-%S')

        # Convert the naive datetime to an aware datetime
        aware_datetime = timezone.make_aware(naive_datetime)

        return aware_datetime
    except (ValueError, IndexError):
        # Return None if the filename doesn't match the expected pattern
        return None


def get_element_by_index(elements, index):
    """
    Safely returns the element at the given index from the list.
    If the index is out of range, returns None.
    """
    return elements[index] if 0 <= index < len(elements) else None


def shift_encode(text, shift=3):
    return ''.join(chr((ord(c) + shift) % 126) for c in text)


def round_to_zero(value, digits):
    if value is None:
        return '-'
    if digits == 0:
        return int(round(value, 0))
    if isinstance(value, Decimal):
        value = float(value)
    return round(value, digits)


class PeriodCalc:

    def __init__(self, period):
        self.period = period

    def _init_statistics(self):
        return {
            _('organiization'): (
                f"{self.period.tenant.name} ({self.period.tenant.code})"),
            _('period'): self.period.name,
            _('start'): self.period.start,
            _('end'): self.period.end
        }

    def _init_count(self):
        return {
            'count': 0,
            'total': 0
        }

    def _excel_total(self, ws, statistics):
        bold = Font(bold=True)

        # Append the total row
        ws.append([
            "Total", '',
            statistics['consumption']['all']['count'],
            statistics['consumption']['all']['total'],
            statistics['consumption']['unit']
        ])

        # Get index of the last row (just added)
        total_row = ws.max_row

        # Apply bold font to the entire row
        for cell in ws[total_row]:
            cell.font = bold

        # Optionally add a spacer row
        ws.append([])

    def _excel_adjust_cols(self, ws):
        ''' Auto-adjust column widths
        '''
        for col in ws.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = (
                max_length + 2
            )

    def create_statistics(self):
        # Init
        statistics = self._init_statistics()
        consumption = {
            'unit': None,
            'areas': {},
            'codes': {},
            'all': self._init_count(),
            'no_value': self._init_count(),
            'total_per_area': {},
            'total_per_code': {},
            'measurements': []
        }

        # Get measurements
        measurements = Measurement.objects.filter(
            route__period=self.period)

        for measurement in measurements:
            # Init
            consumption['all']['count'] += 1
            if measurement.consumption is None:
                consumption['no_value']['count'] += 1
            else:
                consumption['all']['total'] += measurement.consumption

            # Unit
            if not consumption['unit']:
                language = TenantSetup.objects.filter(
                    tenant=self.period.tenant).first().language
                consumption['unit'] = (
                    measurement.counter.category.unit.name.get(language))

            # Adress area
            if measurement.address:
                code = measurement.address.area.code
                consumption['areas'].setdefault(
                    code, measurement.address.area.name)
            else:
                code = _('n/a')
            consumption['total_per_area'].setdefault(code, self._init_count())
            consumption['total_per_area'][code]['count'] += 1
            consumption['total_per_area'][code]['total'] += (
                measurement.consumption or 0)

            # Counter code
            if measurement.counter:
                code = measurement.counter.category.code
                consumption['codes'].setdefault(
                    code, measurement.counter.category.name)
            else:
                code = _('n/a')
            consumption['total_per_code'].setdefault(code, self._init_count())
            consumption['total_per_code'][code]['count'] += 1
            consumption['total_per_code'][code]['total'] += (
                measurement.consumption or 0)

            # Add Measurement
            consumption['measurements'].append({
                'counter_code': measurement.counter.code,
                'date': measurement.datetime.date(),
                'value': measurement.value,
                'consumption': measurement.consumption,
                'route':  f"{measurement.route}",
                'period':  f"{measurement.route.period}"
            })

        statistics.update({
            'consumption': consumption
        })

        return statistics

    def create_excel(self, statistics, filename=None):
        ''' make excel from dict '''
        wb = Workbook()
        ws = wb.active
        ws.title = _("Statistics")
        bold = Font(bold=True)

        # Header info
        unit = statistics['consumption']['unit']
        ws.append(["Organization", statistics['organiization']])
        ws.append(["Period", statistics['period']])
        ws.append(["Start Date", statistics['start']])
        ws.append(["End Date", statistics['end']])
        ws.append([])
        if not filename:
            filename = f"statistics_report_{statistics['period']}.xlsx"

        # Total per area
        headers = [
            _("Code"), _("Description"), _("Count"), _("Total"), _("Unit")
        ]
        ws.append([_("Total per Area")])
        ws.append(headers)
        for code, data in statistics['consumption']['total_per_area'].items():
            try:
                desc = statistics['consumption']['areas'][code]
            except:
                desc = ''
            ws.append([code, desc, data['count'], data['total'], unit])
        self._excel_total(ws, statistics)

        # Total per code
        ws.append([_("Total per Code")])
        ws.append(headers)
        for code, data in statistics['consumption']['total_per_code'].items():
            desc = statistics['consumption']['codes'][code]['de']
            ws.append([code, desc, data['count'], data['total'], unit])
        self._excel_total(ws, statistics)

        # Format headers in bold
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if all(isinstance(cell.value, str) for cell in row if cell.value):
                for cell in row:
                    cell.font = bold

        # Format numbers
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = (
                        CUSTOM_NUMBER_FORMAT if cell.value > 1000
                        else CUSTOM_NUMBER_FORMAT_SMALL)

        # Auto-adjust column widths
        self._excel_adjust_cols(ws)

        # Sheet 2: All Measurements
        ws_measure = wb.create_sheet(title=_("Measurements"))

        # Header row
        headers = [_("Counter Code"), _("Date"), _("Value"), _("Consumption")]
        ws_measure.append(headers)
        for cell in ws_measure[1]:
            cell.font = bold

        # Add measurement rows
        measurements = statistics['consumption'].get('measurements', [])
        for m in measurements:
            ws_measure.append([
                m.get('counter_code'),
                m.get('date'),
                m.get('value'),
                m.get('consumption'),
                m.get('route'),
                m.get('period'),
            ])

        # Auto-adjust column widths
        self._excel_adjust_cols(ws_measure)

        # Save workbook as file
        # wb.save(filename)

        # Save workbook to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the HTTP response
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class RouteManagement:
    '''
    base class to handle Route management
    '''
    def __init__(self, modeladmin, request, route):
        self.modeladmin = modeladmin
        self.request = request
        self.route = route
        self.tenant = route.tenant
        self.created_by = request.user

        # calc
        # start, end
        self.start = self.route.get_start()
        self.end = self.route.get_end()

    def get_last_measurement(self, counter):
        measurement = Measurement.objects.filter(
                tenant=self.tenant,
                counter=counter,
                period=self.route.period
            ).order_by('datetime').last()
        return measurement

    def get_previous_measurement(self, counter):
        measurement = Measurement.objects.filter(
                tenant=self.tenant,
                counter=counter,
                period=self.route.period_previous
            ).order_by('datetime').last()
        return measurement

    def get_comparison_measurements(self, measurement, max=1):
        measurements = Measurement.objects.filter(
                tenant=self.tenant,
                counter=measurement.counter,
                datetime__lt=measurement.datetime
            ).order_by('-datetime')
        return measurements[:max]

    def make_response_json(self, data, filename):
        '''make json file
        input:
            - json data
            - output filename
        return: json file
        '''
        # Create json
        json_data = json.dumps(data, ensure_ascii=False, indent=4)

        # Create the HTTP response and set the appropriate content type with
        # UTF-8 charset
        response = HttpResponse(
            json_data, content_type='application/json; charset=utf-8')

        # Define the file name for the download
        response['Content-Disposition'] = (
            f'attachment; filename="{filename}"')

        return response

    def make_response_excel(self, data_list, queryset, filename):
        '''make excel file
        input:
            - data_list (list of dict)
            - output filename
        return: json file
        '''
        # Init
        headers = tuple(data_list[0].keys())
        data = [tuple(data.values()) for data in data_list]

        # Make
        excel = ExportExcel(
            self.modeladmin, self.request, queryset, filename)
        response = excel.generate_response(
            headers=headers, data=data)

        return response


class RouteCounterExport(RouteManagement):
    '''
    use this class to
    - get_counter_data_json:
        create a json list that gets upload to GFT software

    responsible_user: Brunnenmeister
    key: decipher for JSON export
    '''
    def __init__(
            self, modeladmin, request, route, responsible_user=None,
            route_date=None, energy_type='W', key=None):
        super().__init__(modeladmin, request, route)
        self.username = responsible_user.username if responsible_user else None
        self.name = f'{route.id}, {route.tenant.code}, {route.__str__()}'
        self.route_date = route_date
        self.energy_type = energy_type
        self.key = key  # encryption

    # Helpers
    def data_check(self, show_multiple=False):
        '''Check validity of addresses, counters
        show_multiple: show if a subscriber has multiple counters
        '''
        # Check addresses
        subs_without_address = Subscription.objects.filter(
            tenant=self.tenant,
            address=None,
            start__lte=self.end,
        ).filter(
            Q(end__gte=self.start) | Q(end__isnull=True)
        )
        if subs_without_address:
            messages.warning(
                self.request,
                _("Records having no address: %s") % subs_without_address
            )

        # Check counters
        subs_without_counters = Subscription.objects.annotate(
            counter_count=Count('counters')
        ).filter(
            tenant=self.tenant,
            counter_count=0,
            start__lte=self.end
        ).filter(
            Q(end__gte=self.start) | Q(end__isnull=True)
        )
        if subs_without_counters:
            messages.warning(
                self.request,
                _("Records having no counters: %s") % subs_without_counters
            )

        if show_multiple:
            subs_multiple_counters = Subscription.objects.annotate(
                counter_count=Count('counters')
            ).filter(
                tenant=self.tenant,
                counter_count__gte=1,
                start__lte=self.end
            ).filter(
                Q(end__gte=self.start) | Q(end__isnull=True)
            )
            if subs_without_counters:
                messages.warning(
                    self.request,
                    _("Records having multiple counters: %s") % (
                        subs_multiple_counters)
                )

    def get_addresses(self):
        # Filter Addresses
        if self.route.addresses.exists():
            addresses = self.route.addresses.all()
        else:
            queryset = AddressMunicipal.objects.filter(
                tenant=self.tenant)
            if self.route.areas.exists():
                queryset = queryset.filter(area__in=areas)
            addresses = queryset.all()

        return addresses

    def get_subscriptions(self, addresses):
        # In scope
        queryset = Subscription.objects.filter(
            tenant=self.tenant,
            address__in=addresses,
            start__lte=self.end,
        ).filter(
            Q(end__gte=self.start) | Q(end__isnull=True)
        )

        return queryset.all()

    def make_gft_meter(self, subscription, counter, excel):
        ''' make meter dict for json / excel export
        if excel we optimize data to our needs
        redo !!!
        '''
        # last consumption
        previous_measurement = self.get_previous_measurement(counter)
        if previous_measurement:
            value = previous_measurement.value or 0
            consumption_previous = previous_measurement.consumption or 0
        else:
            value = 0
            consumption_previous = 0

        # Calc min, max
        min, max = value, value
        if consumption_previous:
            min += consumption_previous * float(self.route.confidence_min)
            max += consumption_previous * float(self.route.confidence_max)

        # subscription
        name = subscription.subscriber.__str__()
        if subscription.partner:
            name += ' & ' + subscription.partner.__str__()

        # address
        address = subscription.address
        street = address.stn_label or ''
        nr = address.adr_number or ''

        # make record
        if excel:
            # Make record
            meter = {
                'counter_id': counter.number,
                'address': f"{street} {nr}",
                'subscriber': name,
                'area': address.area,
                'start': start,
                'end': end,
                'consumption_previous': consumption_previous,
                'obiscode': counter.obiscode
            }
        else:
            # current date
            if self.route_date:
                current_date = convert_datetime_to_date(self.route_date)
            else:
                current_date = None

            # previous date
            if self.route.period_previous:
                previous_date = convert_datetime_to_date(
                    self.route.period_previous.end)
            else:
                previous_date = None

            # encryption
            if self.key:
                name = shift_encode(name, self.key)
                nr = shift_encode(street, self.key)
                street = shift_encode(nr, self.key)

            meter = {
                'id': counter.number,
                'energytype': self.energy_type,
                'number': counter.number,
                'hint': json.dumps({
                    'subscription_id': subscription.id,
                    'address_id': address.id,
                    'consumption_previous': round(consumption_previous, 1)
                }),
                'address': {
                    'street': street,
                    'housenr': nr,
                    'city': address.city,
                    'zip': address.zip,
                    'hint': f'address_id: {address.id}',
                },
                'subscriber': {
                    'name': name,
                    'hint': subscription.subscriber.notes
                },
                'value': {
                    'obiscode': counter.category.code,
                    'dateOld': previous_date,
                    'old': round(value, 1),
                    'min': round(min, 1),
                    'max': round(max, 1),
                    'dateCur': current_date
                }
            }

        return meter

    # Methods, main
    def get_counter_data_json(self, excel=False, show_multiple=False):
        '''called to generate the json data for the export to GFT software
        excel: output to excel, creates a much simpler meter data
        show_multiple: show if a subscriber has multiple counters
        '''
        # Check validity of data
        self.data_check(show_multiple)

        # Update route
        data = dict(METER.TEMPLATE)
        data['billing_mde']['route'].update({
            'name': self.name,
            'user': self.username
        })

        # Get subscriptions
        addresses = self.get_addresses()
        subscriptions = self.get_subscriptions(addresses)
        number_of_counters = 0

        # Fill in meters
        for subscription in subscriptions:
            for counter in subscription.counters.all():
                # make meter
                meter = self.make_gft_meter(subscription, counter, excel)
                data['billing_mde']['meter'].append(meter)

                # add route to routes_out in Subscription
                subscription.routes_out.add(self.route)
                number_of_counters += 1

        # Update route
        self.route.number_of_addresses = len(addresses)
        self.route.number_of_subscriptions = len(subscriptions)
        self.route.number_of_counters = number_of_counters
        self.route.status = Route.STATUS.COUNTER_EXPORTED
        self.route.save()

        # return data, if excel only meter records
        return data['billing_mde']['meter'] if excel else data


class RouteCounterExportNew(RouteManagement):
    '''
    use this class to
    - get_counter_data_json:
        create a json list that gets upload to GFT software

    responsible_user: Brunnenmeister
    key: decipher for JSON export
    '''
    def __init__(
            self, modeladmin, request, route, responsible_user=None,
            route_date=None, key=None):
        super().__init__(modeladmin, request, route)
        self.username = responsible_user.username if responsible_user else None
        self.name = f'{route.id}, {route.tenant.code}, {route.__str__()}'
        self.route_date = route_date
        self.key = key  # encryption

    def _get_energy_type(self, counter):
        return 'W'  # only water supported for now

    def _make_gft_meter(self, subscription):
        ''' make meter dict for json / excel export
        if excel we optimize data to our needs
        redo !!!
        '''
        # Init
        counter = subscription.counter

        # last consumption
        previous_measurement = self.get_previous_measurement(counter)
        if previous_measurement:
            value = previous_measurement.value or 0
            consumption_previous = previous_measurement.consumption or 0
        else:
            value = 0
            consumption_previous = 0

        # Calc min, max
        min, max = value, value
        if consumption_previous:
            min += consumption_previous * float(self.route.confidence_min)
            max += consumption_previous * float(self.route.confidence_max)

        # subscription
        name = subscription.subscriber.__str__()
        if subscription.partner:
            name += ' & ' + subscription.partner.__str__()

        # address
        address = subscription.address
        street = address.stn_label or ''
        nr = address.adr_number or ''

        # current date
        if self.route_date:
            current_date = convert_datetime_to_date(self.route_date)
        else:
            current_date = None

        # previous date
        if self.route.period_previous:
            previous_date = convert_datetime_to_date(
                self.route.period_previous.end)
        else:
            previous_date = None

        # encryption
        if self.key:
            name = shift_encode(name, self.key)
            nr = shift_encode(street, self.key)
            street = shift_encode(nr, self.key)

        meter = {
            'id': counter.number,
            'energytype': self._get_energy_type(counter),
            'number': counter.number,
            'hint': json.dumps({
                'subscription_id': subscription.id,
                'address_id': address.id,
                'consumption_previous': round(consumption_previous, 1)
            }),
            'address': {
                'street': street,
                'housenr': nr,
                'city': address.city,
                'zip': address.zip,
                'hint': f'address_id: {address.id}',
            },
            'subscriber': {
                'name': name,
                'hint': subscription.subscriber.notes
            },
            'value': {
                'obiscode': counter.category.code,
                'dateOld': previous_date,
                'old': round(value, 1),
                'min': round(min, 1),
                'max': round(max, 1),
                'dateCur': current_date
            }
        }

        return meter

    def _init_json(self):
        data = dict(METER.TEMPLATE)
        data['billing_mde']['route'].update({
            'name': self.name,
            'user': self.username
        })
        return data

    # Methods, main
    def make_export_file(self, filename, file_type='json'):
        '''called to generate the json data for the export to GFT software
        excel: output to excel, creates a much simpler meter data
        show_multiple: show if a subscriber has multiple counters
        '''
        data = self._init_json()
        count = 0

        # get subscriptions
        if self.route.subscriptions:
            query_subscriptions = self.route.subscriptions
        else:
            query_subscriptions = Subscription.objects.filter(
                tenant=self.tenant, is_inactive=False)

        for subscription in query_subscriptions.all():
            if subscription.counter:
                # Check if measurement already existing
                queryset = Measurement.objects.filter(
                    route=self.route, counter=subscription.counter)
                if queryset.exists():
                    msg = _("subscription {subscription}; already measured.")
                    msg = msg.format(subscription=subscription)
                    messages.warning(self.request, msg)
                else:
                    meter = self._make_gft_meter(subscription)
                    if meter:
                        data['billing_mde']['meter'].append(meter)
                        count += 1
            else:
                messages.warning(
                    self.request, _(f"subscription {subscription}; no counter"))

        # Update route
        # self.route.number_of_addresses = len(addresses)
        self.route.number_of_subscriptions = count  # only include valid ones
        # self.route.number_of_counters = number_of_counters
        self.route.status = Route.STATUS.COUNTER_EXPORTED
        self.route.save()

        messages.info(
            self.request, _(f"json files contains {count} records"))

        # return export_file
        if file_type == 'json':
            return self.make_response_json(data, filename)
        return None


class RouteCounterImport(RouteManagement):
    '''
    use this class to import GFT json list after tour is completed

    this generates for all meters in the json file a measurement,
    unique tenant, counter, route, datetime
    '''
    JSON_MAPPING = {
        # for billing
        'datetime': 'dateKey',  # reference data, e.g. "2025-03-31"
        'value': 'key',   # reference data, use this for billing (default)

        # latest, alternative billing, not used
        'datetime_latest': 'dateCur',
        'value_latest': 'cur',
        'current_battery_level': 'batteryLevel',
    }

    def __init__(self, modeladmin, request, route):
        super().__init__(modeladmin, request, route)

    def create_measurement(self, meter):
        # Get counter
        code = meter['id']
        counter = Device.objects.filter(
            tenant=self.tenant, code=code
        ).first()
        if not counter:
            messages.warning(
                self.request, _(f"counter {code} not found."))
            return None

        # Check data
        value = meter['value']
        if not value:
            messages.warning(
                self.request, _(f"counter {code} has no value."))
            return None

        # Check required keys
        if not all(k in value for k in ['key', 'dateKey', 'cur']):
            messages.warning(
                self.request, _(f"counter {code} has no measurement."))
            return None

        # References
        maintenance = json.loads(meter['hint'])

        # Check address
        address_id = maintenance['address_id']
        address = AddressMunicipal.objects.filter(
            tenant=self.tenant, id=address_id).first()
        if not address:
            messages.warning(
                self.request, _(f"Address {maintenance['address_id']} wrong."))
            return None

        # Check subscription
        subscription_id = maintenance['subscription_id']
        subscription = Subscription.objects.filter(
            tenant=self.tenant, id=subscription_id).first()
        if not subscription:
            messages.warning(
                self.request, _(f"Subscription {subscription_id} wrong."))
            return None

        # Check data
        reference_dt = convert_str_to_datetime(value['dateKey'])
        if not self.start <= reference_dt.date() <= self.end:
            messages.warning(
                self.request,
                _(f"{code}: {reference_dt.date()} not in "
                  f"{self.route.start} to {self.route.end}")
            )
            return None

        # Prepare Measurement data
        data = {
            field_name: value.get(key)
            for field_name, key in self.JSON_MAPPING.items()
        }

        # Convert dates
        for key in ['datetime', 'datetime_latest']:
            data[key] = convert_str_to_datetime(data[key])

        # Calc and update
        data.update({
            # Calc consumption
            'consumption': data['value'] - value['old'],
            'consumption_latest': data['value_latest'] - value['old'],

            # Efficiency analysis
            'address': address,
            'period': self.route.period,
            'subscription': subscription,

            # maintenance
            'created_by': self.created_by,
        })

        # Check if the record has been already imported
        # So it is not possible that same measurement is assigend to two
        # routes
        if Measurement.objects.filter(
                    tenant=self.tenant,
                    counter=counter,
                    datetime=data['datetime']
                ).exists():
            messages.warning(
                self.request, _(f"counter {meter['id']} already measured."))
            return None

        # Store data
        obj = Measurement.objects.create(
            tenant=self.tenant,
            route=self.route,
            counter=counter,
            **data
        )

        return obj

    def process(self, json_file):
        # Load file
        file_data = json_file.read().decode('utf-8')
        data = json.loads(file_data)

        # Check file
        try:
            measurements = data['billing_mde']['meter']
        except:
            raise ValueError(_("Not a valid file"))

        ''' next time
        # Assign route
        route_id = int(data['billing_mde']['route']['name'].split(',')[0])
        self.route = Route.objects.filter(
            tenant=self.tenant, id=route_id).frist()
        if not self.route:
            raise ValueError(_("No valid route id."))
        '''

        # start, end of route
        self.start = self.route.period.start
        self.end = self.route.period.end

        # get meter data
        count = 0
        for meter in measurements:
            measurement = self.create_measurement(meter)
            if measurement:
                count += 1

        # Create an Attachment instance
        attachment = Attachment.objects.create(
            tenant=self.tenant,  # Set the tenant
            content_object=self.route,  # Set the associated route
            file=json_file,  # Uploaded the file
            created_by=self.created_by
        )

        # Add the attachment to the route's attachments
        self.route.attachments.add(attachment)

        # update route
        self.route.status = Route.STATUS.COUNTER_IMPORTED
        self.route.import_file_id = attachment.id
        self.route.save()

        return count


class RouteCounterInvoicing(RouteManagement):
    '''use this to invoice route data

    is_enabled_sync: set true for drafts, set false for others
    '''
    def __init__(
            self, modeladmin, request, route, status, invoice_date,
            is_enabled_sync):
        super().__init__(modeladmin, request, route)
        self.status = status
        self.date = invoice_date
        self.is_enabled_sync = is_enabled_sync

    def _get_quantity(
            self, measurement, article, quantity, rounding_digits,
            days=None):
        ''' quantity, not considered: individual from, to
        '''
        if article.unit.code == 'day' and days:
            # case: days given
            if quantity:
                return days * quantity
        elif quantity:
            # just return quantity
            return quantity
        elif measurement:
            # fill in consumption
            return round(measurement.consumption, rounding_digits)

        # No valid case
        return None  # could not be derived

    def bill(self, subscription, route, check_measurement=True):
        ''' get called from actions '''
        # description
        setup = route.setup
        if subscription.description:
            description = ', ' + subscription.description
        else:
            description = ', ' + setup.description if setup.description else ''

        # billing base
        invoice = {
            'tenant': route.tenant,
            'category': setup.order_category,
            'contract': setup.order_contract,
            'responsible_person': setup.contact,
            # 'dossier': subscription.dossier, must be an invoice but difficult to get
            'date': self.date,
            'status': self.status,
            'is_enabled_sync': self.is_enabled_sync,
            'sync_to_accounting': True,  # immediately sync with cashCtrl
            'created_by': self.created_by
        }

        # associate
        associate = (
            subscription.recipient if subscription.recipient
            else subscription.subscriber
        )
        invoice['associate'] = associate

        # address name
        name = ''
        if associate.title:
            name += primary_language(associate.title.name) + ' '
        if associate.last_name:
            name += f"{associate.first_name} {associate.last_name}"
        if associate.company:
            if name:
                name += ', '
            name += associate.company
        if not subscription.recipient and subscription.partner:
            partner = subscription.partner
            name += '\n'
            if partner.title:
                name += primary_language(partner.title.name) + ' '
            name += f"{partner.first_name} {partner.last_name}"

        # add address
        type, address = associate.get_invoice_address()
        invoice['recipient_address'] = f"{name}\n{address}"

        # building
        if subscription.address:
            building = (
                f"{subscription.address.stn_label} {subscription.address.adr_number}"
            )
        else:
            msg = _("No address for {id}, {subscription}")
            msg = msg.format(
                id=subscription.id,
                subscription=subscription)
            messages.error(self.request, msg)
            return None

        if subscription.address.notes:
            building_notes = ', ' + subscription.address.notes
        else:
            building_notes = ''

        # Get counter_id
        counter_id = subscription.counter.code if subscription.counter else ''

        # Get actual measurement
        measurement = self.get_last_measurement(subscription.counter)
        if measurement:
            # Check consumption
            if measurement.consumption is None:
                msg = _("No consumption for {id}, {subscription}")
                msg = msg.format(
                    id=subscription.id,
                    subscription=subscription)
                messages.error(self.request, msg)
                return None

            # Check invoice
            if measurement.invoice:
                msg = _("{id}, {subscription}: invoice already created for {route}.")
                msg = msg.format(
                    id=subscription.id,
                    subscription=subscription,
                    route=route)
                messages.error(self.request, msg)
                return None

            # check day vs. period
            unit_code = (
                'day' if (
                    route.start
                    or route.end
                    or subscription.start > measurement.period.start
                    or (subscription.end
                        and subscription.end < measurement.period.end)
                ) else 'period'
            )
            if unit_code == 'day':
                # start
                start = max(subscription.start or self.start, self.start)
                end = max(subscription.end or self.end, self.end)
                days = (end - start).days + 1
            else:
                days = None

            # Get comparison consumption
            value_new = round_to_zero(measurement.value, setup.rounding_digits)
            value_old = '-'
            comparisons = self.get_comparison_measurements(
                measurement, max=1)
            if comparisons:
                comparison = comparisons[0]
                value_old = round_to_zero(
                    comparison.value, setup.rounding_digits)
            else:
                comparison = None
                msg = _("No comparison for {id}, {subscription}")
                msg = msg.format(
                    id=subscription.id,
                    subscription=subscription)
                messages.warning(self.request, msg)

            if comparison and comparison.consumption:
                consumption = round_to_zero(
                    comparison.consumption, setup.rounding_digits)
            else:
                consumption = ''
        else:
            # default: raise Error message
            if check_measurement:
                msg = _("No actual measurement for {id}, {subscription}")
                msg = msg.format(
                    id=subscription.id,
                    subscription=subscription)
                messages.error(self.request, msg)
                return None

            # check day vs. period
            unit_code = 'day' if (route.start or route.end) else 'period'
            if unit_code == 'day':
                # start
                start = max(subscription.start or self.start, self.start)              
                end = min(subscription.end or self.end, self.end)
                days = (end - start).days + 1
            else:
                days = None

            # consumption
            consumption, value_new, value_old = '-', '-', '-'

        # subscriber_short_name
        if subscription.recipient:
            # Invoice recipient if not subscriber.
            subscriber_short_name = f", {subscription.subscriber.short_name}"
            if subscription.partner:
                subscriber_short_name += (
                    f" und {subscription.partner.short_name}")
        else:
            subscriber_short_name = ''

        # header, use SafeDict to avoid error of variable not in template
        template = setup.header
        invoice['header'] = template.format_map(SafeDict(
            building=building,
            building_notes=building_notes,
            description=description,
            subscription_id=f"S-{subscription.id}",
            subscriber_short_name=subscriber_short_name,
            start=format_date(self.start),
            end=format_date(self.end),
            consumption=consumption,
            counter_id=counter_id,
            counter_new=value_new,
            counter_old=value_old
        ))

        # description
        invoice['description'] = (
            f"{route.name}, {building or '-'}{building_notes}, {description}")
        if subscription.tag:
            invoice['description'] += ', ' + subscription.tag

        # create article items
        items = []
        sub_articles = SubscriptionArticle.objects.filter(
            subscription=subscription
        ).order_by('article__nr')
        for subscription_article in sub_articles:
            article = subscription_article.article
            quantity = subscription_article.quantity

            if unit_code == 'day' and article.unit.code == 'period':
                # Replace article by daily
                article = Article.objects.filter(
                    tenant=article.tenant,
                    nr=article.nr + ARTICLE_NR_POSTFIX_DAY,
                ).first()

            quantity = self._get_quantity(
                measurement, article, quantity, setup.rounding_digits, days)
            if quantity is None:
                msg = _("{subscription}: no valid measurement for {article}.")
                msg = msg.format(subscription=subscription, article=article)
                messages.error(self.request, msg)
                return None

            items.append(dict(
                tenant=route.tenant,
                article=article,
                quantity=self._get_quantity(
                    measurement, article, quantity, setup.rounding_digits,
                    days),
                created_by=self.created_by
            ))

        # No error, create objects
        # Create invoice
        invoice_obj = OutgoingOrder.objects.create(**invoice)

        # Create items
        try:
            for item in items:
                item['order'] = invoice_obj
                obj = OutgoingItem.objects.create(**item)
        except Exception as e:
            print("Error creating OutgoingItem:", e)
            invoice_obj.delete()

        # Update measurement, should be True except "Pauschal"
        if measurement:
            measurement.invoice = invoice_obj
            measurement.save()
        else:
            msg = _("{subscription}: no measurement.")
            messages.warning(
                self.request, msg.format(subscription=subscription))

        return invoice_obj


class MeasurementAnalyse:
    '''
    select measurements and analyse
    measurements is queryset from admin.py
    '''
    def __init__(self, modeladmin, measurements):
        self._model = modeladmin.model
        self.queryset = queryset

    def analyse(self):
        '''
        analyse data and returns dictionary
        '''
        # Distinct values of periods (assuming 'route' relates to periods)
        distinct_periods = set([x.route.period for x in self.queryset.all()])

        # Distinct values of routes
        distinct_routes = set([x.route for x in self.queryset.all()])

        # Distinct values of areas (assuming 'address__categories' link to areas)
        distinct_areas = Area.objects.filter(
            id__in=self.queryset.values('address__area')
        ).distinct()

        # Start and end of periods
        start = min([x.period.start for x in self.queryset.all()])
        end = max([x.period.end for x in self.queryset.all()])

        # Min and Max datetime
        min_date = self.queryset.aggregate(
            Min('datetime'))['datetime__min'].date()
        max_date = self.queryset.aggregate(
            Max('datetime'))['datetime__max'].date()

        # Sum of consumption
        consumption = 0
        for measurement in self.queryset:
            if measurement.consumption:
                consumption += measurement.consumption

        # Growth
        consumption_change = (
            1- (total_consumption - total_consumption_previous)
            / total_consumption_previous
        ) if total_consumption_previous else None

        return {
            'periods': distinct_periods,
            'routes': distinct_routes,
            'areas': distinct_areas,
            'start': start,
            'end': end,
            'min_date': min_date,
            'max_date': max_date,
            'consumption': total_consumption,
            'consumption_previous': total_consumption_previous,
            'consumption_change': consumption_change
        }

    def output_excel(self, data, filename=None, title=None):
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title if title else _('Consumption Analysis')

        if not filename:
            filename = 'output.xlsx'

        # Add header with bold font
        header = [_('Label'), _('Value')]
        ws.append(header)

        bold_font = Font(bold=True)
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = bold_font  # Make header bold

        # Add data rows
        rows = [
            [_('Records Processed'), data['record_count']],
            [_('Periods'), ', '.join(x.name for x in data['periods'])],
            [_('Routes'), ', '.join(x.name for x in data['routes'])],
            [_('Areas'), ', '.join(x.code for x in data['areas'])],
            [_('Period Start, End'), f"{data['start']} - {data['end']}"],
            [_('Value Dates From - To'), f"{data['min_date']} - {data['max_date']}"],
            [_('Total Consumption'), f"{data['consumption']:.0f} m³"],
            [_('Previous'),
             f"{data['consumption_previous']:.0f}"
             if data['consumption_previous'] else '-'],
            [_('Change in %'), data['consumption_change']],
        ]

        for row in rows:
            ws.append(row)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2  # Padding

        # Prepare HTTP response
        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            ),
        )
        response['Content-Disposition'] = f"attachment; filename={filename}"

        # Save workbook to response
        wb.save(response)
        return response