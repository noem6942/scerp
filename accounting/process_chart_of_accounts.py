# process_chart_of_accounts.py
from django.utils.translation import gettext_lazy as _
from enum import Enum
from openpyxl import load_workbook

# Definitions
class TYPE(Enum):
    BALANCE = 1  # Bilanz
    FUNCTIONAL = 2  # Funktionale Gliederung
    INCOME = 3  # Erfolgsrechnung
    INVEST = 5  # Investitionsrechnung


HEADERS = [
    'account_number', 'account_4_plus_2', 'name', 'notes', 
    'hrm_1', 'description_hrm_1'
]


# mixins
class Account(object):
    '''parse account table and do consistency checks
    input variables:
        file_path: excel file
        type_: TYPE
        function: None if excel is functional, else int

    output variables:
        account: string for sorting        
    '''

    def __init__(self, file_path, type_, function=None):
        # Init
        self.file_path = file_path
        self.type = self.get_valid_type(type_)
        
        # function
        if self.is_valid_function(function):
            self.function = function            

    # init functions
    def get_valid_type(self, type_):
        return next(x for x in TYPE if x.value == type_ or x == type_)

    def is_valid_function(self, function):
        '''
        Validate the function based on its type.

        :param type_: The type of account (e.g., FUNCTIONAL, BALANCE).
        :param function: The function to validate.
        :raises ValueError: If the function is invalid based on the type.
        '''
        if self.type in [TYPE.FUNCTIONAL, TYPE.BALANCE]:
            # If type is FUNCTIONAL or BALANCE, function should not be provided
            if function is not None:
                raise ValueError(_('no valid function given'))

        # Check if function is provided and within the valid range        
        if (function is not None 
                and not 0 <= function <= MAX_ACCOUNT_NR[TYPE.FUNCTIONAL]):
            raise ValueError(_('no valid function given'))

        return True  # Return True if valid


    # excel functions
    def read_excel_file(self, file_path):
        '''read an excel sheet and interprete EVERY cell as string.
            i.e. empty cell -> ''
                 111.11 -> '111.11'
                 012 -> '012'
        '''
        # Load the workbook
        wb = load_workbook(filename=file_path, data_only=False)  # data_only=False to get formulas too
        ws = wb.active  # Use the active sheet
        
        # Iterate through the rows in the worksheet
        rows = []
        for row in ws.iter_rows(values_only=True):
            # Convert each cell to string while keeping leading zeros
            string_row = [
                str(cell).strip() if cell is not None else '' 
                for cell in row]
            rows.append(string_row)
        
        return rows

    # parse    
    def get_accounts(self):
        # Get content
        headers = []
        rows = self.read_excel_file(self.file_path)
        
        # Get accounts
        accounts = []
        for row_nr, row in enumerate(rows, start=1):
            # Get headers (first non-empty row that contains 'Sachkonto')
            if not headers:
                if row[0] == 'Sachkonto':
                    headers = HEADERS  
                continue

            # Analyze
            data = dict(zip(headers, row))            
            data['row_nr'] = row_nr
            
            # Account, category 
            if data['account_number'] or data['account_4_plus_2']:
                accounts.append(data)

        return accounts
