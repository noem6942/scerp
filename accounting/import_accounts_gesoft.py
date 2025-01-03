# import_accounts_gesoft.py
from .models import ACCOUNT_TYPE, CATEGORY_HRM

from openpyxl import load_workbook


class ACCOUNT_SIDE:
    # Used for assignment
    EXPENSE = 1
    REVENUE = 2 
    CLOSING = 9


class Import(object):

    def __init__(self, file_path, account_type, category_hrm=None):
        # Account
        self.account_type = account_type
        self.category_hrm = category_hrm
        self.scope = CATEGORY_HRM.get_scope(category_hrm)  # e.g. [4, 6]
        
        # read content
        workbook = load_workbook(file_path)

        # Get a specific sheet
        sheet = workbook.active  # Default to the first sheet or use workbook['SheetName']
        
        # Iterate over rows and columns
        # `values_only=False ensures we get formats as well (leading zeroes)
        self.rows = [row for row in sheet.iter_rows(values_only=False)] 

    def format_with_leading_zeros(self, value, format_string):
        """
        Formats an integer or string to match the leading zero pattern in the format string.
        
        :param value: The value to be formatted (int or str).
        :param format_string: The format string defining the desired pattern (e.g., '0000').
        :return: A formatted string matching the pattern in `format_string`
            if possible otherwise value
        """
        # Determine the length of the format string
        length = len(format_string)

        # Convert value to an integer, then to a string with leading zeros
        if type(value) == int:
            # do not skip leading zeros
            return f"{int(value):0{length}d}"
        else:
            return str(value)

    def get_accounts(self):
        # Init
        accounts = []
        function = None
        row_count = len(self.rows)        
        count = 0
        
        # Parse
        for nr, row in enumerate(self.rows):  
            # Get cell_0
            cell_0 = row[0]            

            # Check if valid position            
            if type(cell_0.value) not in [int, float]:
                continue            
            
            # Get name, is_category
            name = row[1].value
            is_category = False if '0.00' in cell_0.number_format else True            
                        
            # Get account_number + leading zeros for first col
            account_number = cell_0.value
            account_number_str = self.format_with_leading_zeros(
                cell_0.value, cell_0.number_format)
            
            # check if broken name in next row
            if nr < row_count - 1:
                next_row = [x.value for x in self.rows[nr + 1]]
                if next_row[0] is None and next_row[1]:
                    name += ' ' + next_row[1]
            
            # Function, account_number_str
            if is_category:
                # re-assign function
                function = account_number_str
            else:
                # clean account_number_str               
                # Convert to a float                
                numeric_value = float(account_number_str)                
                # Format as a float with two decimal places                
                account_number_str = f"{numeric_value:.2f}"  

            # Read numbers
            if self.account_type == ACCOUNT_TYPE.BALANCE:
                # Assign balance, budget, previous
                budget = None
                (previous, _add, _subtract, balance
                ) = [None] * 4 if is_category else [x.value for x in row[2:]]
            else:
                (income, expense,
                 budget_income, budget_expense,
                 previous_income, previous_expense
                ) = [None] * 6 if is_category else [x.value for x in row[2:]]
            
                # Assign balance, budget, previous
                balance = income if income is not None else expense
                budget = (
                    budget_income if budget_income is not None 
                    else budget_expense)
                previous = (
                    previous_income if previous_income is not None 
                    else previous_expense)            
                 
            # Make account
            accounts.append({
                'function': function,  # never None, for sorting
                'account_type': self.account_type,
                'is_category': is_category,
                'account_number': account_number_str,
                'name': name,
                'balance': balance,
                'budget': budget,
                'previous': previous,
            })
            
        if self.account_type == ACCOUNT_TYPE.INVEST:
            # Add missing
            accounts.append({
                'function': '0',  
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '0',
                'name': 'Allgemeine Verwaltung',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '02', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '02',
                'name': 'Allgemeine Dienste',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '02', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '026',
                'name': 'Bürgergemeinde',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '7', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '7',
                'name': 'Umweltschutz und Raumordnung',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '71', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '71',
                'name': 'Wasserversorgung SF',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '710', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '710',
                'name': 'Wasserversorgung Spezialfinanzierung (SF)',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '8', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '8',
                'name': 'Volkswirtschaft',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '82', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '82',
                'name': 'Forstwirtschaft',
                'balance': None,
                'budget': None,
                'previous': None
            })
            accounts.append({
                'function': '820', 
                'account_type': ACCOUNT_TYPE.INVEST,
                'is_category': True,
                'account_number': '820',
                'name': 'Forstwirtschaft',
                'balance': None,
                'budget': None,
                'previous': None
            })
            
        return accounts
        