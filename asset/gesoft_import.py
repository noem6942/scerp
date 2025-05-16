'''
asset/gesoft_import.py

Run this before billing/gesoft_import.py!

'''
import json
import logging
from openpyxl import load_workbook
from pathlib import Path

from django.conf import settings

from core.models import Tenant
from scerp.mixins import parse_gesoft_to_datetime
from .models import DEVICE_STATUS, AssetCategory, Device, EventLog


logger = logging.getLogger(__name__)

DEVICE_IS_ENABLED_SYNC = False  # do not sync devices, yet
DATE_NULL = '01.01.1900'

HOTWATER_COUNTER_IDS = [
    19871932,
    19871931,
    19871929,
    19871934,
    19871933,
    19871930
]


class OBIS_CODE:
    WATER = '8-0:1.0.0'
    HOT_WATER = '9-0:1.0.0'


class ImportDevice:

    STATUS_MATCHING = {
        'Lager': DEVICE_STATUS.IN_STOCK,
        'Montiert': DEVICE_STATUS.MOUNTED,
        'Weggeworfen': DEVICE_STATUS.DISPOSED,
    }

    def __init__(self, tenant_id):
        '''
            id of tenant
        '''
        # From setup
        self.tenant = Tenant.objects.get(id=tenant_id)
        self.created_by = self.tenant.created_by
        self.asset_category_query = AssetCategory.objects.filter(
            tenant=self.tenant)

    def load(self, file_name):
        '''get Addresses
        we use this to get the invoicing addresses

        file_name = 'Abonnenten Gebühren einzeilig.xlsx'
        writes address_data with abo_nr as key
        '''
        file_path = Path(
            settings.BASE_DIR) / 'asset' / 'fixtures' / file_name
        wb = load_workbook(file_path)
        ws = wb.active  # Or wb['SheetName']
        rows = [row for row in ws.iter_rows(values_only=True)]

        # Read
        for row_nr, row in enumerate(rows, start=1):
            cells = row
            if cells[0] and isinstance(cells[0], (int, float)) and cells[2]:
                # Get data
                (werk_nr, _, zaehlerart, zw, jg, _, st, _, volt, zoll_mm, _,
                 typ, eichdatum, status, *_) = row

                # dt
                if eichdatum:
                    dt = parse_gesoft_to_datetime(eichdatum)
                elif jg:
                    dt = parse_gesoft_to_datetime(jg)
                else:
                    dt = parse_gesoft_to_datetime(DATE_NULL)

                try:
                    status = self.STATUS_MATCHING[status.strip()]
                except:
                    raise ValueError(f"{row_nr}: {status} not found")

                logger.info(f"parsing #{row_nr} {werk_nr}")

                # get asset_category
                if float(werk_nr) in HOTWATER_COUNTER_IDS:
                    asset_category = self.asset_category_query.filter(
                        code=OBIS_CODE.HOT_WATER).first()
                else:
                    asset_category = self.asset_category_query.filter(
                        code=OBIS_CODE.WATER).first()
                if not asset_category:
                    raise ValueError(f"no asset category found for {werk_nr}")

                # edit or update counter
                counter, created = Device.objects.update_or_create(
                    tenant=self.tenant,
                    category=asset_category,
                    code=werk_nr,
                    defaults=dict(
                        created_by=self.created_by,
                        date_added=dt.date(),
                        status=status,
                        date_disposed=(
                            dt.date() if status==DEVICE_STATUS.DISPOSED
                            else None),
                        number=werk_nr,
                        is_enabled_sync=DEVICE_IS_ENABLED_SYNC
                    )
                )
                logger.info(f"storing counter {werk_nr}, created: {created}")

                # edit or update event, no defaults
                event, _created = EventLog.objects.get_or_create(
                    tenant=self.tenant,
                    device=counter,
                    datetime=dt,
                    status=status,
                    defaults=dict(created_by=self.created_by)
                )
                logger.info(f"storing {status} {werk_nr}, created: {created}")

                if eichdatum:
                    status = DEVICE_STATUS.CALIBRATED
                    event, _created = EventLog.objects.get_or_create(
                        tenant=self.tenant,
                        device=counter,
                        datetime=dt.replace(hour=12, minute=0, second=0),
                        status=status,
                        defaults=dict(created_by=self.created_by)
                    )
                    logger.info(f"storing {status} {werk_nr}, created: {created}")


def update_counter_assets(tenant_id):
    # introduce factor
    codes = [OBIS_CODE.WATER, OBIS_CODE.HOT_WATER]

    # update factor
    categories = AssetCategory.objects.filter(
        tenant__id=tenant_id,
        code__in=codes,
        counter_factor=1
    ).all()
    logger.info(f"categories {categories}")

    # update, do not trigger event in cashCtrl
    categories.update(counter_factor=1)

    # make negative counters
    for category in categories:
        obj = category
        obj.pk = None
        obj.counter_factor = -1
        obj.name = {k: v + ' neg.' for k,v in obj.name.items()}
        obj.save()
        logger.info(f"saved {obj}")

    # assign
    category_old = AssetCategory.objects.filter(
        tenant__id=tenant_id,
        code=OBIS_CODE.WATER,
        counter_factor=1
    ).first()
    category_new = AssetCategory.objects.filter(
        tenant__id=tenant_id,
        code=OBIS_CODE.WATER,
        counter_factor=-1
    ).first()

    devices = Device.objects.filter(
        tenant__id=tenant_id,
        category=category_old,
        code__contains='.'
    )
    for device in devices:
        device.category = category_new
        device.sync_to_accounting = True
        device.save()
        logger.info(f"saved {device}")
