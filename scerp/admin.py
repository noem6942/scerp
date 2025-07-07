'''
scerp/admin.py

Admin configuration for the scerp app.

Helpers for admin.py

'''
import json
import openpyxl
import re
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from datetime import datetime, date
from decimal import Decimal

from django.conf import settings
from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.forms import Textarea
from django.http import HttpResponse
from django.db.models import TextChoices
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.utils.encoding import force_str
from django.utils.formats import date_format
from django.utils.translation import get_language, gettext_lazy as _

from .admin_base import FIELDSET, TenantFilteringAdmin, RelatedModelInline
from .mixins import primary_language


class PAGE_ORIENTATION(TextChoices):
    LANDSCAPE = 'landscape', _('Landscape')
    PORTRAIT = 'portrait', _('Portrait')


# format
def format_big_number(value, thousand_separator=None, round_digits=None):
    '''
    use settings.THOUSAND_SEPARATOR and 2 commas for big numbers
    value: float
    '''
    if value is None:
        return None

    # Round
    if round_digits:
        value = round(value, round_digits)
    print("*value", value)
    # Format number
    if thousand_separator is None:
        thousand_separator = settings.THOUSAND_SEPARATOR
    return f'{value:,.2f}'.replace(',', thousand_separator)


def format_percent(value, precision):
    '''
    use settings.THOUSAND_SEPARATOR and 2 commas for big numbers
    value: float
    '''
    if value is None:
        return None

    # Format number
    return f'{value:,.{precision}f}%'


def get_help_text(model, field_name):
    '''
    Get help text from model
    '''
    # pylint: disable=W0212
    return model._meta.get_field(field_name).help_text


def make_language_fields(field_name):
    # Currently show only LANGUAGE_CODE_PRIMARY
    languages = [
        (lang_code, lang) for lang_code, lang in settings.LANGUAGES
        if lang_code == settings.LANGUAGE_CODE_PRIMARY
    ]

    # Check if only show preferred languagen    
    return [
        f'{field_name}_{lang_code}'
        for lang_code, _lang in languages
    ]


def is_required_field(model, field_name):
    '''
    Check if required field
    '''
    # pylint: disable=W0212
    field = model._meta.get_field(field_name)
    return not field.blank and not field.null


def verbose_name(model):
    '''
    Get verbose_name from model
    '''
    # pylint: disable=W0212
    return model._meta.verbose_name


def verbose_name_plural(model):
    '''
    Get verbose_name from model
    '''
    # pylint: disable=W0212
    return model._meta.verbose_name_plural


def verbose_name_field(model, field_name):
    '''
    Get verbose_name from model field
    '''
    # pylint: disable=W0212
    return model._meta.get_field(field_name).verbose_name


def is_html(string):
    return bool(re.search(r'<[a-z][\s\S]*>', string, re.IGNORECASE))


def html_to_number(html_string):
    ''' convert html -> float e.g.
        html_string = '<span style="text-align: right; display: block;">4&#x27;238.00</span>'
        returns: 4238.00
    '''
    # Extract the text content
    soup = BeautifulSoup(html_string, 'html.parser')
    number_str = soup.text.strip()

    # Remove the thousands separator (apostrophe)
    number_str = number_str.replace("'", "")

    # Convert to float
    try:
        return float(number_str)
    except:
        raise ValueError('not a number')


class Display:

    def boolean(value):
        return '✔' if value else '✘'

    def datetime(value, default='-'):
        '''
        Display date time nice
        '''
        if value is None:
            return default
        return date_format(value, format='DATETIME_FORMAT')

    def align_right(value):
        html = '<span style="text-align: right; display: block;">{}</span>'
        return format_html(html, value)

    def big_number(value, round_digits=None, thousand_separator=None):
        '''
        use settings.THOUSAND_SEPARATOR and 2 commas for big numberss
        '''
        if value is None:
            return None

        # Format number
        try:
            number_str = format_big_number(
                value,
                round_digits=round_digits,
                thousand_separator=thousand_separator)
        except:
            number_str = value
        html = '<span style="text-align: right; display: block;">{}</span>'
        return format_html(html, number_str)

    def percentage(value, precision=1):
        '''
        show percenate with{ precision} commas
        '''
        if value is None:
            return None

        # Format number
        number_str = format_percent(value, precision)
        html = '<span style="text-align: right; display: block;">{}</span>'
        return format_html(html, number_str)

    def hierarchy(level, name):
        '''Function to print out hierarchy names nice;
            add spaces before the string if is_category == False
        '''
        if level == 1:
            return format_html(f'<b>{name.upper()}</b>')
        if level == 2:
            return format_html(f'<b>{name}</b>')
        return format_html(f'<i>{name}</i>')

    def json(value, sort=False):
        '''
        Print string for json
        '''
        if not value:
            return ''

        try:
            # Format JSON data with indentation and render it as preformatted text
            formatted_json = json.dumps(value, indent=4, ensure_ascii=False)
            return format_html(
                '<pre style="font-family: monospace;">{}</pre>',
                formatted_json)
        except ValueError as e:
            return f'Value Error displaying data: {e}'
        except (KeyError, TypeError) as e:  # Catch specific exceptions
            return f'Key or Type Error displaying data: {e}'
        except Exception as e:  # pylint: disable=W0718
            # Last resort: Catch any other exceptions
            return f'Unexpected error displaying data: {e}'

    def link(url, name, target='_blank'):
        '''
        Generates a clickable link for the Django admin interface.

        Args:
            url (str): The URL to link to.
            name (str): The text to display for the link.

        Returns:
            str: HTML string with a clickable link.
        '''
        return format_html(
            '<a href="{}" target="{}">{}</a>', url, target, name)

    def list(items):
        output_list = [f'<li>{item}</li>' for item in items]
        return format_html(''.join(output_list))

    def photo(url_field):
        '''
        Display photo
        '''
        if url_field:
            return mark_safe(
                f'<img src="{url_field.url}" width="60" height="60" '
                f'style="object-fit: cover;" />')
        return ''

    def verbose_name(def_cls, field):
        '''
        Display verbose name from field
        '''
        f_cls = getattr(def_cls, 'Field')
        return getattr(f_cls, field)['verbose_name']

    def name_w_levels(obj):
        '''obj needs to have name, is_category, level
        '''
        name = format_name(obj)
        if obj.is_category:
            level = obj.level
            if level < 3:
                return format_hierarchy(obj.level, name)
        return name


class Export:

    def __init__(self, modeladmin, request, queryset, filename):
        self.modeladmin = modeladmin
        self.request = request
        self.queryset = queryset
        self.filename = filename

    def make_headers(self, headers, data=[]):
        # Convert headers if necessary
        headers = [
            x if isinstance(x, str) else force_str(x)
            for x in headers
        ]

        # Write Column Headers
        if not headers and data:
            headers = [f'col{i}' for i, _ in enumerate(data[0], start=1)]

        return headers

    def clean_value(self, value):
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        elif isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, str) and is_html(value):
            try:
                # Look if it's a number
                return html_to_number(value)
            except:
                return value
        elif isinstance(value, (int, float)):
            return value
        return force_str(value)

    def get_data(self):
        ''' get data from modeladmin queryset if not specified in modeladmin
        '''
        # Check if existing
        func = getattr(self.modeladmin, 'get_data', None)
        if func:
            data_list = func(self.request, self.queryset)

        # Create list
        else:
            data_list = []
            for item in self.queryset.all():
                item_list = []
                for fieldname in self.modeladmin.list_display:
                    field, *sub_fields = fieldname.split('__')
                    if len(sub_fields) > 1:
                        raise ValueError("Can only output one foreign key")
                    elif sub_fields:
                        value_parent = getattr(item, field)
                        value = getattr(value_parent, sub_fields[0])
                    elif fieldname == 'display_is_inactive':
                        value = item.is_inactive
                    elif fieldname == 'display_is_protected':
                        value = item.is_protected
                    elif fieldname.startswith('display_notes'):
                        value = item.notes
                    elif fieldname.startswith('display_attachment'):
                        value = item.attachments.exists()
                    elif fieldname.startswith('display'):
                        value = getattr(self.modeladmin, fieldname)(item)
                    else:
                        value = getattr(item, field)
                    item_list.append(self.clean_value(value))
                data_list.append(item_list)

        return data_list

    def get_headers(self):
        ''' get headers from modeladmin queryset if not specified in modeladmin
        '''
        # Check if existing
        func = getattr(self.modeladmin, 'get_headers', None)
        if func:
            return func(self.request, self.queryset)

        # Create list
        header_list = []
        for fieldname in self.modeladmin.list_display:
            if fieldname == 'display_attachment_icon':
                value = _('attachments')
            elif fieldname.startswith('display_'):
                value = fieldname.replace('display_', '')
            # elif fieldname.startswith('display'):
            #     value_parent = getattr(self.modeladmin, fieldname)
            #     value = getattr(value_parent, 'short_description')
            elif '__' in fieldname:
                value = fieldname
            else:
                value = verbose_name_field(self.modeladmin.model, fieldname)
            header_list.append(self.clean_value(value))

        return header_list


class ExportExcel(Export):
    '''
    Generates an Excel file from provided data with customizable headers,
    footers, and column widths.
    '''
    def __init__(
            self, modeladmin, request, queryset, filename='output.xlsx',
            ws_title='Exported Data', header={}, footer={}, orientation=None):
        '''
        Args:
            filename (str): Name of the exported file.
            title (str): worksheet title
            header (dict)
            footer (dict)
        '''
        super().__init__(modeladmin, request, queryset, filename)
        self.ws_title = ws_title
        self.header = header
        self.footer = footer
        self.orientation = (
            orientation if orientation else PAGE_ORIENTATION.LANDSCAPE)

    def set_layout(self, ws):
        # A4 Page Setup
        # Set to A4 paper size
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # Landscape mode for better readability
        if self.orientation == PAGE_ORIENTATION.LANDSCAPE:
            ws.page_setup.orientation = PAGE_ORIENTATION.LANDSCAPE
        elif self.orientation == PAGE_ORIENTATION.PORTRAIT:
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        else:
            raise ValueError(_("No valid page orientation"))

        # Center the print horizontally
        ws.print_options.horizontalCentered = True

        # Center the print vertically
        ws.print_options.verticalCentered = True
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        # Set Header/Footer
        ws.oddHeader.left.text = self.header.get('left', '')
        ws.oddHeader.center.text = self.header.get('center', '')
        ws.oddHeader.right.text = self.header.get('right', '')

        ws.oddFooter.left.text = self.footer.get('left', '')
        ws.oddFooter.center.text = self.footer.get('center', '')
        ws.oddFooter.right.text = self.footer.get('right', '')

    def generate_response(self, col_widths=None, headers=[], data=[]):
        '''
        Generates an Excel file from provided data with customizable headers, footers, and column widths.

        Args:
            data (list of lists):
                Rows of data, where each row is a list of values.
            headers (list): retrieved from modeladmin (default)
            data (list): retrieved from queryset (default)
            col_widths (list, optional):
            List of column widths; defaults to auto-adjust.

        Returns:
            HttpResponse: Excel file as an HTTP response.
        '''
        # Create a new workbook & worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.ws_title
        self.set_layout(ws)

        # Convert headers if necessary
        headers = headers if headers else self.get_headers()
        ws.append(headers)

        # Make headers bold
        bold_font = Font(bold=True)
        for col_num, _ in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).font = bold_font

        # Get data and write in Rows
        data = data if data else self.get_data()
        for row in data:
            ws.append(row)

        # Adjust Column Widths
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            if col_widths and len(col_widths) >= col_num:
                ws.column_dimensions[col_letter].width = (
                    col_widths[col_num - 1])
            else:
                ws.column_dimensions[col_letter].width = (
                    max(len(str(header)) + 2, 12))

        # Prepare HTTP Response
        response = HttpResponse(
            content_type=(
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet')
        )
        response['Content-Disposition'] = (
            f"attachment; filename={self.filename}")
        wb.save(response)

        return response


class ExportJSON(Export):
    '''
    Generates an JSON file from provided data
    '''
    def __init__(self, modeladmin, request, queryset, filename='output.json'):
        '''
        Args:
            filename (str): Name of the exported file. Will
        '''
        super().__init__(modeladmin, request, queryset, filename)

    def generate_response(self):
        '''
        Generates an JSON file from provided data
        Args:
            data (list of lists):
                Rows of data, where each row is a list of values.

        Returns:
            HttpResponse: JSON file as an HTTP response.
        '''
        # Convert headers if necessary
        headers = self.get_headers()

        # Clean data and make list of dicts
        data_list = self.get_data()
        data = [dict(zip(headers, x)) for x in data_list]

        # Create json
        json_data = json.dumps(data, ensure_ascii=False)

        # Create the HTTP response and set the appropriate content type with
        # UTF-8 charset
        response = HttpResponse(
            json_data, content_type='application/json; charset=utf-8')

        # Define the file name for the download (you can customize it as needed)
        response['Content-Disposition'] = (
            f'attachment; filename="{self.filename}"')

        return response


# Decorators
class BaseAdmin:
    '''
    basic class
    all security filtering is done in TenantFilteringAdmin
    '''

    @admin.display(description=_('Name'))
    def display_name(self, obj):
        try:
            return primary_language(obj.name)
        except:
            return ' '

    @admin.display(description=_('Number'))
    def display_number(self, obj):
        return Display.big_number(obj.number)

    @admin.display(description='')
    def display_is_inactive(self, obj):
        return '🔒' if obj.is_inactive else ' '

    @admin.display(description='')
    def display_is_protected(self, obj):
        return '🔒' if obj.is_protected else ' '

    @admin.display(description=_('Photo'))
    def display_photo(self, obj):
        return Display.photo(obj.photo)

    @admin.display(description='Fi')
    def display_attachment_icon(self, obj):
        '''Displays a paperclip 📎 or folder 📂 icon if attachments exist.'''
        if obj.attachments.exists():  # ✅ Efficient query
            url = obj.get_attachment_link(nr=1)
            link = f'<a href="{url}" target="_blank">📎</a>'
            return mark_safe(link)
        return ' '  # No icon if no attachments'

    @admin.display(description='')
    def display_notes(self, obj):
        '''Displays a hint (tooltip) with the note text if available.'''
        if obj.notes:
            return format_html(
                '<span style="cursor: pointer; border-bottom: 1px dotted #555;" '
                'title="{}">📝</span>', obj.notes
            )  # ✅ Cursor + underline effect
        return ' '

    @admin.display(description=_('last update'))
    def display_last_update(self, obj):
        return obj.modified_at

    @admin.display(description=_('Name'))
    def display_name_singular(self, obj):
        try:
            return primary_language(obj.name_singular)
        except:
            return ''

    @admin.display(description=_('Name Plural'))
    def display_name_plural(self, obj):
        try:
            return primary_language(obj.name_plural)
        except:
            return ''

    @admin.display(description=_('Parent'))
    def display_parent(self, obj):
        return self.display_name(obj.parent)

    @admin.display(description=_('Wording in document'))
    def display_document_name(self, obj):
        return primary_language(obj.document_name)

    @admin.display(description=_('Percentage'))
    def display_percentage(self, obj):
        return Display.percentage(obj.percentage, 1)

    @admin.display(description=_('Percentage Flat'))
    def display_percentage_flat(self, obj):
        return Display.percentage(obj.percentage_flat, 1)

    @admin.display(description=_('Balance'))
    def display_link_to_company(self, person):
        if not person.company:
            return '-'  # Fallback if company is missing
        url = f'../person/{person.id}/'
        return format_html('<a href="{}">{}</a>', url, person.company)

    @admin.display(description=_('Category'))
    def display_category_type(self, obj):
        return obj.category.get_type_display()

    @admin.display(description=_('description'))
    def display_description(self, obj):
        return primary_language(obj.description)

    # used in accounting ---------------------------------------
    @admin.display(description=_('last update'))
    def display_last_update(self, obj):
        return obj.modified_at

    @admin.display(description=_('Name Plural'))
    def display_name_plural(self, obj):
        try:
            return primary_language(obj.name_plural)
        except:
            return ''

    @admin.display(description=_('Parent'))
    def display_parent(self, obj):
        return self.display_name(obj.parent)

    @admin.display(description=_('Balance'))
    def display_link_to_company(self, person):
        if not person.company:
            return "-"  # Fallback if company is missing
        url = f"../person/{person.id}/"
        return format_html('<a href="{}">{}</a>', url, person.company)

    @admin.display(
        description=_('function'))
    def display_function(self, obj):
        return obj.account_number if obj.is_category else ' '

    @admin.display(description=_('position nr.'))
    def position_number(self, obj):
        return ' ' if obj.is_category else obj.account_number

    @admin.display(description=_('actual +'))
    def display_end_amount_credit(self, obj):
        if obj.category_hrm in (CATEGORY_HRM.EXPENSE, CATEGORY_HRM.ASSET):
            balance = 0 if obj.end_amount is None else obj.end_amount
            return Display.big_number(balance)
        return ' '

    @admin.display(description=_('actual -'))
    def display_end_amount_debit(self, obj):
        if obj.category_hrm in (CATEGORY_HRM.REVENUE, CATEGORY_HRM.LIABILITY):
            balance = 0 if obj.end_amount is None else obj.end_amount
            return Display.big_number(balance)
        return ' '

    @admin.display(description=_('balance +'))
    def display_balance_credit(self, obj):
        if obj.category_hrm in (CATEGORY_HRM.EXPENSE, CATEGORY_HRM.ASSET):
            balance = 0 if obj.balance is None else obj.balance
            return Display.big_number(balance)
        return ' '

    @admin.display(description=_('balance -'))
    def display_balance_debit(self, obj):
        if obj.category_hrm in (CATEGORY_HRM.REVENUE, CATEGORY_HRM.LIABILITY):
            balance = 0 if obj.balance is None else obj.balance
            return Display.big_number(balance)
        return ' '

    @admin.display(description=_('balance'))
    def display_balance(self, obj):
        balance = 0 if obj.balance is None else obj.balance
        return Display.big_number(balance)

    @admin.display(description=_('budget'))
    def display_budget(self, obj):
        return Display.big_number(obj.budget)

    @admin.display(description=_('previous'))
    def display_previous(self, obj):
        return Display.big_number(obj.previous)

    @admin.display(description=_('cashCtrl'))
    def display_cashctrl(self, obj):
        if obj.c_id or obj.c_rev_id:
            return '🪙'  # (Coin): \U0001FA99
        return ' '
    

class BaseTabularInline(RelatedModelInline):
    ''' 
    same concept as BaseAdmin, currently no methods 
    '''
    pass
