"""
scerp/mixins.py
"""
import logging
import os
import pyproj  # A library for coordinate transformations
import secrets
import string
import yaml
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from django.apps import apps
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from django.utils.timezone import make_aware
from django.conf import settings
from django.utils.translation import activate, get_language, gettext


logger = logging.getLogger(__name__)  # Using the app name for logging

# helpers, use this for all models in all apps
current_timezone = timezone.get_current_timezone()

# Get User.
# This approach will work even if you later decide to switch to a custom
# user model, as get_user_model() always fetches the model defined in
# AUTH_USER_MODEL.
User = get_user_model()


# I/O functions
def read_excel(file_path, header_nr=1, string_cols=[]):
    """
    Reads an Excel file and returns a list of rows as dictionaries.
    Columns listed in `string_cols` will be coerced to strings (to preserve leading zeros, etc).
    
    :param file_path: Path to the .xlsx file
    :param header_nr: Row number containing headers (1-based)
    :param string_cols: List of column names to force as strings
    :return: List of dictionaries, one per row
    """
    wb = load_workbook(file_path, data_only=False)
    ws = wb.active

    # Read header row
    header_row = ws[header_nr]
    headers = [cell.value for cell in header_row]

    data = []
    for row in ws.iter_rows(min_row=header_nr + 1, values_only=False):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers):
                continue  # skip extra columns

            col_name = headers[col_idx]
            value = cell.value

            # Force to string if in string_cols
            if col_name in string_cols:
                if value is None:
                    value = ''
                elif isinstance(value, float) and value.is_integer():
                    value = str(int(value))
                else:
                    value = str(value)

            row_data[col_name] = value

        data.append(row_data)

    return data


def read_yaml_file(app_name, filename_yaml):
    '''
    Load the YAML file with app_name as parent dir
    '''
    file_path = os.path.join(
        settings.BASE_DIR, app_name, filename_yaml)
    with open(file_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)


# A function to convert Swiss coordinates to WGS84 (latitude/longitude)
def convert_ch1903_to_wgs84(easting, northing):
    try:
        # LV95 to WGS84
        transformer = pyproj.Transformer.from_crs(
            'EPSG:2056', 'EPSG:4326', always_xy=True)
        lon, lat = transformer.transform(easting, northing)
        return lat, lon
    except Exception as e:
        raise ValidationError(f"Error converting coordinates: {e}")


def format_date(date, format='%d.%m.%Y'):
    formatted_date = date.strftime(format)
    return formatted_date


# Models
def get_admin():
    '''
    Handsome for assigning base logging data to have the admin user
    '''
    return User.objects.get(username='admin')


def get_code_w_name(instance):
    code = instance.code + ' ' if instance.code else ''
    return f"{code} {primary_language(instance.name)}"


def is_url_friendly(name):
    '''ensure right org_names are chosen
    '''
    slugified_name = slugify(name)
    # Check if the slugified name is non-empty and identical to the original, lowercased name
    return slugified_name == name.lower() and bool(slugified_name)

    
def show_hidden(key):
    return '*' * len(key)  


def generate_random_password(length=settings.PASSWORD_LENGTH):
    """Generate a secure random password."""
    alphabet = string.ascii_letters + string.digits + string.punctuation
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def get_translations(text):
    ''' return language map for text if existing
    does not work! shows always german
    '''
    name = {}
    current_language = get_language()  # Store the currently active language

    for lang_code, lang_name in settings.LANGUAGES:
        activate(lang_code)  # Temporarily activate the language
        name[lang_code] = gettext(text)  # get translation

    # Restore the original language
    activate(current_language)

    return name


def make_multi_language(name, language=None):
    '''
    Creates a dictionary with language codes as keys and initializes all
    values as None, except for the primary or specified language, which
    is assigned the given name.

    Args:
        name (str): The value to assign to the primary or specified language.
        language (str, optional): The language code to assign `name` to.
                                  Defaults to the primary language from settings.

    Returns:
        dict: A dictionary where keys are language codes, and values are either
              None or the provided name for the selected language.

    Example:
        >>> make_multi_language('Hello')
        {'en': 'Hello', 'de': None, 'fr': None, 'es': None}

        >>> make_multi_language('Hallo', 'de')
        {'en': None, 'de': 'Hallo', 'fr': None, 'es': None}
    '''
    value = {lang: None for lang, _ in settings.LANGUAGES}
    if not language:
        language = settings.LANGUAGE_CODE_PRIMARY
    value[language] = name
    return value


def make_timeaware(naive_datetime):
    '''used to save datetimes from external data
    '''
    return timezone.make_aware(naive_datetime, current_timezone)


def parse_gesoft_to_datetime(date_input):
    """Convert a date string to a timezone-aware Django datetime object."""
    if not date_input:
        return None
    elif isinstance(date_input, datetime):
        # Already a datetime, just make it timezone-aware
        return make_aware(date_input)
    elif isinstance(date_input, int):
        date_input = str(date_input)  # only year given

    formats = ["%Y-%m-%d %H:%M:%S", "%d.%m.%Y", "%Y"]  # Supported formats

    for fmt in formats:
        try:
            dt = datetime.strptime(date_input, fmt)  # Parse string to datetime
            return make_aware(dt)  # Automatically applies Django's timezone
        except ValueError:
            continue  # Try the next format

    raise ValueError(f"Unsupported date format: {date_input}")  # Handle errors


def primary_language(value_dict):
    '''show language default instead of all values
    '''
    if value_dict is None or isinstance(value_dict, str):
        return value_dict

    # get languages
    try:
        language = get_language().split('-')[0]
    except:
        language = settings.LANGUAGE_CODE_PRIMARY

    # Check if None
    if value_dict:
        values = dict(value_dict)
    else:
        values = {lang: None for lang, _ in settings.LANGUAGES}

    if values and language in values:
        return values[language]
    elif values and settings.LANGUAGE_CODE_PRIMARY in values:
        return values[language]
    return str(value_dict)


# signals, load and init yaml data
def init_yaml_data(
        app_name, tenant, created_by, filename_yaml, accounting_setup=None,
        model_filter=[]):

    return  # currently no installations when create a new tenant

    # Load the YAML file
    file_path = os.path.join(
        settings.BASE_DIR, app_name, filename_yaml)
    with open(file_path, 'r', encoding='utf-8') as file:
        data = yaml.safe_load(file)

    # Create data
    try:
        with transaction.atomic():
            for model_name, data_list in data.items():
                if model_filter and model_name not in model_filter:
                    continue
                for data in data_list:
                    # Prepare data
                    data.update({
                        'tenant': tenant,
                        'created_by': created_by
                    })
                    if accounting_setup:
                        data['setup'] = accounting_setup

                    # Create data
                    model = apps.get_model(
                        app_label=app_name, model_name=model_name)
                    obj = model.objects.create(**data)

                    logger.info(f"{model}: created {obj}.")

        # At this point, all operations are successful and committed
        logger.info("All operations completed successfully!")
    except Exception as e:
        # If any exception occurs, the transaction is rolled back
        logger.error(f"Transaction failed: {e}")

        # Re-raise the exception to propagate it further
        raise  # This will raise the same exception that was caught
